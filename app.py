from __future__ import annotations

import html
import json
import random
import re
import time
import uuid
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

from backend import backend_enabled, post_backend
from reporting import build_version_csv, build_versions_pdf, encode_attachment

st.set_page_config(
    page_title="Coal-to-Clean Jurisdictional Readiness Index 2026",
    page_icon=Path(__file__).parent / "assets" / "favicon.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

DATA_FILE = Path(__file__).parent / "data" / "Index_Model_Streamlit.xlsx"
SHEET_NAME = "Streamlit"
OVERLAY_SHEET_NAME = "OVERLAY"
MARKET_DRIVERS_SHEET = "MARKET_DRIVERS"
MARKET_CONTEXT_SHEET = "MARKET_CONTEXT"
MARKET_DEVELOPMENTS_SHEET = "MARKET_DEVELOPMENTS"


PILLAR_SOURCES = [
    "PILLAR 1 score\nEnergy system conditions",
    "PILLAR 2 score\nPolicy and transition commitment",
    "PILLAR 3 score\nGovernance and institutional capacity",
    "PILLAR 4 score\nCarbon market maturity",
    "PILLAR 5 score\nMacro-financial conditions",
    "PILLAR 6 score\nJust transition and social credibility",
]

PILLAR_FULL = {
    PILLAR_SOURCES[0]: "Pillar 1: Energy system conditions",
    PILLAR_SOURCES[1]: "Pillar 2: Policy and transition commitment",
    PILLAR_SOURCES[2]: "Pillar 3: Governance and institutional capacity",
    PILLAR_SOURCES[3]: "Pillar 4: Carbon market maturity",
    PILLAR_SOURCES[4]: "Pillar 5: Macro-financial conditions",
    PILLAR_SOURCES[5]: "Pillar 6: Just transition and social credibility",
}

PILLAR_TABLE = {
    PILLAR_SOURCES[0]: "Pillar: Energy system conditions",
    PILLAR_SOURCES[1]: "Pillar: Policy and transition commitment",
    PILLAR_SOURCES[2]: "Pillar: Governance and institutional capacity",
    PILLAR_SOURCES[3]: "Pillar: Carbon market maturity",
    PILLAR_SOURCES[4]: "Pillar: Macro-financial conditions",
    PILLAR_SOURCES[5]: "Pillar: Just transition and social credibility",
}

PILLAR_SHORT = [
    "Energy system conditions",
    "Policy and transition commitment",
    "Governance and institutional capacity",
    "Carbon market maturity",
    "Macro-financial conditions",
    "Just transition and social credibility",
]

PILLAR_DESCRIPTIONS = [
    "Signals whether the power system can replace coal while maintaining reliable electricity supply.",
    "Signals whether announced transition policies, coal commitments and regulatory measures are likely to be sustained and implemented.",
    "Signals whether institutions, regulation and public administration can support credible project delivery.",
    "Signals experience in developing carbon projects, issuing credits, transferring units and reaching buyers.",
    "Signals whether financing, currency and payment conditions support investable and bankable projects.",
    "Signals whether worker, community and social-protection risks are recognised and managed during coal phase-down.",
]

BASE_WEIGHTS = [0.20, 0.20, 0.15, 0.20, 0.15, 0.10]

RECOMMENDATIONS = {
    PILLAR_SOURCES[0]: "Test coal-retirement feasibility, replacement capacity and grid-readiness assumptions at project level.",
    PILLAR_SOURCES[1]: "Test whether policy commitments translate into durable approvals, implementation and credible retirement pathways.",
    PILLAR_SOURCES[2]: "Strengthen regulatory, institutional and counterparty safeguards before committing capital or delivery obligations.",
    PILLAR_SOURCES[3]: "Verify project-development capability, issuance delivery, transfer infrastructure and evidence of buyer use.",
    PILLAR_SOURCES[4]: "Assess currency, financing and payment risks, including suitable risk-mitigation and capital structures.",
    PILLAR_SOURCES[5]: "Examine worker, community and social-protection arrangements within the transition plan.",
}

INDUSTRY_MULTIPLIERS = {
    "Agriculture": [0.95, 1.00, 1.00, 1.10, 1.00, 1.10],
    "Aviation": [0.95, 1.05, 1.00, 1.25, 1.05, 0.95],
    "Automotive": [1.05, 1.05, 1.00, 1.00, 1.10, 1.00],
    "Banking and financial services": [1.00, 1.05, 1.10, 1.10, 1.25, 0.90],
    "Carbon markets and climate services": [0.90, 1.10, 1.05, 1.35, 1.00, 0.90],
    "Cement and building materials": [1.15, 1.10, 1.00, 0.90, 1.15, 0.90],
    "Chemicals": [1.10, 1.05, 1.00, 0.95, 1.15, 0.95],
    "Construction": [1.10, 1.05, 1.00, 0.95, 1.10, 1.00],
    "Consumer goods": [0.95, 1.00, 1.00, 1.15, 1.00, 1.05],
    "Data centres": [1.10, 1.00, 1.00, 1.10, 1.00, 0.95],
    "Development finance and multilateral institutions": [1.00, 1.15, 1.20, 1.05, 1.20, 1.10],
    "Food and beverage": [0.95, 1.00, 1.00, 1.10, 1.00, 1.10],
    "Healthcare": [0.95, 1.00, 1.00, 1.05, 1.00, 1.05],
    "Hospitality and tourism": [0.95, 1.00, 1.00, 1.15, 1.00, 1.05],
    "Infrastructure": [1.10, 1.05, 1.00, 0.95, 1.10, 1.00],
    "Insurance": [0.95, 1.05, 1.10, 1.05, 1.25, 1.00],
    "Manufacturing": [1.10, 1.05, 1.00, 0.95, 1.15, 0.95],
    "Metals and steel": [1.15, 1.05, 1.00, 0.95, 1.15, 0.95],
    "Mining": [1.15, 1.05, 1.00, 0.95, 1.15, 0.95],
    "Mobility and fleet services": [1.05, 1.00, 1.00, 1.00, 1.10, 1.00],
    "Oil and gas": [1.15, 1.05, 1.00, 0.95, 1.10, 1.00],
    "Pharmaceuticals": [0.95, 1.00, 1.00, 1.05, 1.00, 1.05],
    "Power generation": [1.20, 1.10, 1.00, 0.90, 1.00, 1.10],
    "Professional services": [1.00, 1.05, 1.10, 1.05, 1.05, 1.00],
    "Public transport and rail": [1.10, 1.05, 1.00, 0.95, 1.10, 1.05],
    "Pulp and paper": [1.10, 1.05, 1.00, 0.95, 1.10, 1.00],
    "Real estate": [1.05, 1.05, 1.00, 0.95, 1.10, 1.00],
    "Retail": [0.95, 1.00, 1.00, 1.10, 1.00, 1.05],
    "Road freight and logistics": [1.00, 1.05, 1.00, 1.05, 1.10, 0.95],
    "Semiconductors and electronics": [1.05, 1.00, 1.00, 1.05, 1.10, 0.95],
    "Shipping": [1.00, 1.05, 1.00, 1.15, 1.10, 0.95],
    "Technology": [1.05, 1.00, 1.00, 1.10, 1.00, 0.95],
    "Telecommunications": [1.05, 1.00, 1.00, 1.05, 1.00, 1.00],
    "Textiles and apparel": [0.95, 1.00, 1.00, 1.10, 1.00, 1.05],
    "Utilities and energy networks": [1.15, 1.10, 1.00, 0.90, 1.00, 1.10],
    "Waste management": [1.00, 1.00, 1.00, 1.10, 1.00, 1.05],
    "Other": [1.00, 1.00, 1.00, 1.00, 1.00, 1.00],
}

ROLE_OPTIONS = [
    "Corporate buyer or end user",
    "Market intermediary or adviser",
    "Investor or lender",
    "Project developer or originator",
    "Government or regulator",
    "Standard, registry or assurance provider",
    "Civil society, research or academia",
    "Asset owner or operator",
    "Other",
]

SURVEY_ITEMS = [
    {
        "key": "reliable_power",
        "label": "Reliable power during the transition",
        "description": "Keeping electricity dependable while coal is replaced by renewable or other lower-emissions power.",
        "pillars": [0],
    },
    {
        "key": "government_follow_through",
        "label": "Government follow-through",
        "description": "The host government implements and maintains its announced energy-transition policies and regulations.",
        "pillars": [1],
    },
    {
        "key": "project_delivery",
        "label": "Ability to deliver the project",
        "description": "Coal-to-clean projects can move from planning through construction, operation and credit issuance.",
        "pillars": [2],
    },
    {
        "key": "carbon_market_track_record",
        "label": "Carbon-market track record",
        "description": "The host jurisdiction has credible experience developing projects, issuing credits and reaching buyers.",
        "pillars": [3],
    },
    {
        "key": "stable_finance",
        "label": "Stable finance and payments",
        "description": "The host jurisdiction offers workable financing, currency and payment conditions for project delivery.",
        "pillars": [4],
    },
    {
        "key": "workers_communities",
        "label": "Protection for workers and communities",
        "description": "The transition addresses impacts on workers, communities and local livelihoods affected by coal phase-down.",
        "pillars": [5],
    },
    {
        "key": "time_to_delivery",
        "label": "Time to delivery",
        "description": "Projects can reach operation and issue credits within a predictable timeframe.",
        "pillars": [1, 2],
    },
    {
        "key": "credit_price",
        "label": "Credit price",
        "description": "The expected carbon-credit price is competitive enough to influence your decision.",
        "pillars": [4],
    },
]

VALIDATION_SURVEY_VERSION = "2026-07-29-v5"
SURVEY_ROUTE_PARAM = "r"
SURVEY_ROUTE_TOKEN = "c2c26-7k4m9"
VALIDATION_SCALE = {
    1: "Not important",
    2: "Slightly important",
    3: "Moderately important",
    4: "Very important",
    5: "Critical",
}

VALIDATION_FACTORS: list[dict[str, Any]] = [
    {
        "question_id": "F01",
        "key": "renewable_growth",
        "label": "Renewable growth",
        "description": "Solar and/or wind generation has grown steadily in the jurisdiction.",
        "construct_id": "C05B",
        "construct": "Observed renewables-growth credibility",
        "indicator_id": "P2-I5",
    },
    {
        "question_id": "F02",
        "key": "reliable_replacement_power",
        "label": "Reliable replacement power",
        "description": "New renewable generation and grid improvements can replace the coal plant without materially weakening electricity reliability.",
        "construct_id": "X01",
        "construct": "Grid connection and reliability deliverability",
        "indicator_id": "Context only",
    },
    {
        "question_id": "F03",
        "key": "renewable_policy_support",
        "label": "Renewable policy support",
        "description": "The government has clear renewable-energy policies and incentives, supported by active procurement, auctions or other programmes.",
        "construct_id": "C06",
        "construct": "Renewable-policy implementation",
        "indicator_id": "P2-I1",
    },
    {
        "question_id": "F04",
        "key": "net_zero_commitment",
        "label": "Net-zero commitment",
        "description": "The jurisdiction has a clear and credible net-zero target.",
        "construct_id": "C03",
        "construct": "Policy durability and implementation credibility",
        "indicator_id": "P2-I8",
    },
    {
        "question_id": "F05",
        "key": "coal_phaseout_deadline",
        "label": "Coal phase-out deadline",
        "description": "The jurisdiction has set a clear target year for ending unabated coal power.",
        "construct_id": "C03",
        "construct": "Policy durability and implementation credibility",
        "indicator_id": "P2-I8",
    },
    {
        "question_id": "F06",
        "key": "government_follow_through",
        "label": "Government follow-through",
        "description": "The government backs its coal-transition commitments with laws, funding and visible implementation.",
        "construct_id": "C03",
        "construct": "Policy durability and implementation credibility",
        "indicator_id": "P2-I8",
    },
    {
        "question_id": "F07",
        "key": "coal_subsidy_reform",
        "label": "Coal subsidy reform",
        "description": "The government has removed or is reducing subsidies and other support that keep coal power artificially competitive.",
        "construct_id": "C06",
        "construct": "Renewable-policy implementation",
        "indicator_id": "P2-I1",
    },
    {
        "question_id": "F08",
        "key": "no_new_coal_approvals",
        "label": "No new coal approvals",
        "description": "The government has stopped approving or permitting new coal-fired power plants.",
        "construct_id": "C04A",
        "construct": "New coal pipeline contradiction",
        "indicator_id": "P2-I2",
    },
    {
        "question_id": "F09",
        "key": "no_new_coal_construction",
        "label": "No new coal construction",
        "description": "New coal-fired power plants are not entering or remaining under construction.",
        "construct_id": "C04B",
        "construct": "Coal construction contradiction",
        "indicator_id": "P2-I3",
    },
    {
        "question_id": "F10",
        "key": "no_coal_restarts",
        "label": "No coal restarts",
        "description": "Mothballed or previously closed coal-fired power plants are not brought back into operation.",
        "construct_id": "C03",
        "construct": "Policy durability and implementation credibility",
        "indicator_id": "P2-I8",
    },
    {
        "question_id": "F11",
        "key": "carbon_pricing_framework",
        "label": "Carbon-pricing framework",
        "description": "The jurisdiction has an operating carbon tax or emissions trading system covering major emitters.",
        "construct_id": "C07",
        "construct": "Carbon-pricing relevance to retirement economics",
        "indicator_id": "P2-I7",
    },
    {
        "question_id": "F12",
        "key": "meaningful_coal_carbon_cost",
        "label": "Meaningful coal carbon cost",
        "description": "Coal-fired power generation faces a carbon cost high enough to influence operating or investment decisions.",
        "construct_id": "C07",
        "construct": "Carbon-pricing relevance to retirement economics",
        "indicator_id": "P2-I7",
    },
    {
        "question_id": "F13",
        "key": "carbon_credit_track_record",
        "label": "Carbon-credit track record",
        "description": "The jurisdiction has hosted carbon-credit projects – such as REDD+, methane capture, clean cooking or biochar – that have progressed to credit issuance or buyer use.",
        "construct_id": "C12A/C12B/C12C",
        "construct": "Carbon-project development, issuance and buyer-use track record",
        "indicator_id": "P4-I1/P4-I2/P4-I3",
    },
    {
        "question_id": "F14",
        "key": "international_carbon_market_procedures",
        "label": "International carbon-market procedures",
        "description": "The jurisdiction has clear procedures for approving and accounting for carbon credits transferred internationally.",
        "construct_id": "C11A",
        "construct": "Article 6 institutional readiness",
        "indicator_id": "P4-I4",
    },
    {
        "question_id": "F15",
        "key": "international_transfer_readiness",
        "label": "Article 6 readiness",
        "description": "The jurisdiction is putting Article 6 procedures in place – for example, designating a responsible authority, publishing approval and authorisation procedures, establishing or accessing registry infrastructure, entering bilateral cooperation arrangements, or approving activities and authorising credits for international transfer.",
        "construct_id": "C11B",
        "construct": "Article 6 operational execution",
        "indicator_id": "P4-I5",
    },
    {
        "question_id": "F16",
        "key": "public_sector_coordination",
        "label": "Public-sector coordination",
        "description": "Ministries, regulators, utilities and other public agencies can coordinate the approvals, financing and implementation needed for a complex coal-retirement transaction.",
        "construct_id": "C08",
        "construct": "Public-sector execution capacity",
        "indicator_id": "P3-I2",
    },
    {
        "question_id": "F17",
        "key": "predictable_approvals",
        "label": "Predictable approvals",
        "description": "Energy regulations, licences and approval processes are clear and remain reasonably stable.",
        "construct_id": "C09A",
        "construct": "Regulatory predictability",
        "indicator_id": "P3-I3",
    },
    {
        "question_id": "F18",
        "key": "enforceable_agreements",
        "label": "Enforceable agreements",
        "description": "Power-purchase agreements, licences, contracts and government commitments are likely to be honoured and enforced.",
        "construct_id": "C09B",
        "construct": "Contract and licence enforceability",
        "indicator_id": "P3-I4",
    },
    {
        "question_id": "F19",
        "key": "local_transition_finance",
        "label": "Local transition finance",
        "description": "Local banks, investors and capital markets can help finance the plant’s early closure, renewable replacement and related transition measures.",
        "construct_id": "C16",
        "construct": "Domestic transition-finance capacity",
        "indicator_id": "P5-I1",
    },
    {
        "question_id": "F20",
        "key": "currency_payment_stability",
        "label": "Currency and payment stability",
        "description": "Exchange rates, currency convertibility, hedging and payment conditions can support long-term project finance.",
        "construct_id": "C15",
        "construct": "Currency and convertibility risk",
        "indicator_id": "P5-I4",
    },
    {
        "question_id": "F21",
        "key": "social_security_protections",
        "label": "Social security and protections",
        "description": "The jurisdiction has social-protection systems – such as unemployment support, pensions, healthcare or disability support – that can protect people affected by economic disruption.",
        "construct_id": "C19",
        "construct": "Social protection and community resilience",
        "indicator_id": "P6-I2",
    },
    {
        "question_id": "F22",
        "key": "worker_transition_pathways",
        "label": "Worker transition pathways",
        "description": "Workers losing coal-related employment can access realistic retraining, new jobs and local economic opportunities.",
        "construct_id": "C20",
        "construct": "Labour-market absorption and reskilling capacity",
        "indicator_id": "P6-I3",
    },
]

SURVEY_COUNTRIES = [
    'Afghanistan',
    'Albania',
    'Algeria',
    'American Samoa',
    'Andorra',
    'Angola',
    'Anguilla',
    'Antarctica',
    'Antigua and Barbuda',
    'Argentina',
    'Armenia',
    'Aruba',
    'Australia',
    'Austria',
    'Azerbaijan',
    'Bahamas',
    'Bahrain',
    'Bangladesh',
    'Barbados',
    'Belarus',
    'Belgium',
    'Belize',
    'Benin',
    'Bermuda',
    'Bhutan',
    'Bolivia',
    'Bosnia and Herzegovina',
    'Botswana',
    'Bouvet Island',
    'Brazil',
    'British Indian Ocean Territory',
    'British Virgin Islands',
    'Brunei',
    'Bulgaria',
    'Burkina Faso',
    'Burundi',
    'Cabo Verde',
    'Cambodia',
    'Cameroon',
    'Canada',
    'Caribbean Netherlands',
    'Cayman Islands',
    'Central African Republic',
    'Chad',
    'Chile',
    'China',
    'Christmas Island',
    'Cocos (Keeling) Islands',
    'Colombia',
    'Comoros',
    'Cook Islands',
    'Costa Rica',
    'Croatia',
    'Cuba',
    'Curaçao',
    'Cyprus',
    'Czech Republic',
    'Côte d’Ivoire',
    'Denmark',
    'Djibouti',
    'Dominica',
    'Dominican Republic',
    'DR Congo',
    'Ecuador',
    'Egypt',
    'El Salvador',
    'Equatorial Guinea',
    'Eritrea',
    'Estonia',
    'Eswatini',
    'Ethiopia',
    'Falkland Islands',
    'Faroe Islands',
    'Fiji',
    'Finland',
    'France',
    'French Guiana',
    'French Polynesia',
    'French Southern Territories',
    'Gabon',
    'Gambia',
    'Georgia',
    'Germany',
    'Ghana',
    'Gibraltar',
    'Greece',
    'Greenland',
    'Grenada',
    'Guadeloupe',
    'Guam',
    'Guatemala',
    'Guernsey',
    'Guinea',
    'Guinea-Bissau',
    'Guyana',
    'Haiti',
    'Heard Island and McDonald Islands',
    'Honduras',
    'Hong Kong',
    'Hungary',
    'Iceland',
    'India',
    'Indonesia',
    'Iran',
    'Iraq',
    'Ireland',
    'Isle of Man',
    'Israel',
    'Italy',
    'Jamaica',
    'Japan',
    'Jersey',
    'Jordan',
    'Kazakhstan',
    'Kenya',
    'Kiribati',
    'Kosovo',
    'Kuwait',
    'Kyrgyzstan',
    'Laos',
    'Latvia',
    'Lebanon',
    'Lesotho',
    'Liberia',
    'Libya',
    'Liechtenstein',
    'Lithuania',
    'Luxembourg',
    'Macao',
    'Madagascar',
    'Malawi',
    'Malaysia',
    'Maldives',
    'Mali',
    'Malta',
    'Marshall Islands',
    'Martinique',
    'Mauritania',
    'Mauritius',
    'Mayotte',
    'Mexico',
    'Micronesia',
    'Moldova',
    'Monaco',
    'Mongolia',
    'Montenegro',
    'Montserrat',
    'Morocco',
    'Mozambique',
    'Myanmar',
    'Namibia',
    'Nauru',
    'Nepal',
    'Netherlands',
    'New Caledonia',
    'New Zealand',
    'Nicaragua',
    'Niger',
    'Nigeria',
    'Niue',
    'Norfolk Island',
    'North Korea',
    'North Macedonia',
    'Northern Mariana Islands',
    'Norway',
    'Oman',
    'Pakistan',
    'Palau',
    'Palestine',
    'Panama',
    'Papua New Guinea',
    'Paraguay',
    'Peru',
    'Philippines',
    'Pitcairn',
    'Poland',
    'Portugal',
    'Puerto Rico',
    'Qatar',
    'Republic of the Congo',
    'Romania',
    'Russia',
    'Rwanda',
    'Réunion',
    'Saint Barthélemy',
    'Saint Helena, Ascension and Tristan da Cunha',
    'Saint Kitts and Nevis',
    'Saint Lucia',
    'Saint Martin',
    'Saint Pierre and Miquelon',
    'Saint Vincent and the Grenadines',
    'Samoa',
    'San Marino',
    'Sao Tome and Principe',
    'Saudi Arabia',
    'Senegal',
    'Serbia',
    'Seychelles',
    'Sierra Leone',
    'Singapore',
    'Sint Maarten',
    'Slovakia',
    'Slovenia',
    'Solomon Islands',
    'Somalia',
    'South Africa',
    'South Georgia and the South Sandwich Islands',
    'South Korea',
    'South Sudan',
    'Spain',
    'Sri Lanka',
    'Sudan',
    'Suriname',
    'Svalbard and Jan Mayen',
    'Sweden',
    'Switzerland',
    'Syria',
    'Taiwan',
    'Tajikistan',
    'Tanzania',
    'Thailand',
    'Timor-Leste',
    'Togo',
    'Tokelau',
    'Tonga',
    'Trinidad and Tobago',
    'Tunisia',
    'Turkmenistan',
    'Turks and Caicos Islands',
    'Tuvalu',
    'Türkiye',
    'Uganda',
    'Ukraine',
    'United Arab Emirates',
    'United Kingdom',
    'United States',
    'United States Minor Outlying Islands',
    'Uruguay',
    'US Virgin Islands',
    'Uzbekistan',
    'Vanuatu',
    'Vatican City',
    'Venezuela',
    'Vietnam',
    'Wallis and Futuna',
    'Western Sahara',
    'Yemen',
    'Zambia',
    'Zimbabwe',
    'Åland Islands',
]


COUNTRY_GROUP_SEPARATOR = "────────── Other countries ──────────"
ASEAN_PRIORITY_COUNTRIES = [
    "Singapore",
    "Brunei",
    "Cambodia",
    "Indonesia",
    "Laos",
    "Malaysia",
    "Myanmar",
    "Philippines",
    "Thailand",
    "Timor-Leste",
    "Vietnam",
]
ASEAN_PRIORITY_SET = set(ASEAN_PRIORITY_COUNTRIES)
OTHER_SURVEY_COUNTRIES = sorted(
    (country for country in SURVEY_COUNTRIES if country not in ASEAN_PRIORITY_SET),
    key=str.casefold,
)
SURVEY_COUNTRY_OPTIONS = [
    *ASEAN_PRIORITY_COUNTRIES,
    COUNTRY_GROUP_SEPARATOR,
    *OTHER_SURVEY_COUNTRIES,
]


def clear_country_separator_from_selectbox(state_key: str) -> None:
    """Return a country selectbox to its placeholder if the divider is clicked."""
    if st.session_state.get(state_key) == COUNTRY_GROUP_SEPARATOR:
        st.session_state[state_key] = "Select a country"


def clear_country_separator_from_multiselect(state_key: str) -> None:
    """Prevent the visual divider from being retained as a selected market."""
    selected = list(st.session_state.get(state_key, []))
    if COUNTRY_GROUP_SEPARATOR in selected:
        st.session_state[state_key] = [
            value for value in selected if value != COUNTRY_GROUP_SEPARATOR
        ]

COAL_PHASEOUT_INITIATIVES = [
    "Just Energy Transition Partnerships",
    "Asian Development Bank Energy Transition Mechanism",
    "Coal to Clean Credit Initiative",
    "Verra VM0052 transition-credit methodology",
    "Gold Standard JUST: Coal Decommissioning methodology",
    "Transition Credits Coalition, or TRACTION",
    "Other",
    "I do not know enough about these initiatives to rank them",
]

REQUIRED_COLUMNS = [
    "Country",
    "ISO3",
    "Eligible",
    "Base index rank",
    "Base index score",
    *PILLAR_SOURCES,
]

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap');

:root {
    --main-surface: rgba(236, 242, 249, .55);
    --sidebar: #9fbedf;
    --ink: rgba(0, 0, 0, .90);
    --ink-muted: rgba(0, 0, 0, .62);
    --accent: #2b5688;
    --accent-mid: #79a6d2;
    --accent-soft: #8bb0da;
    --accent-pale: #c4d7ed;
    --card: rgba(255, 255, 255, .94);
    --card-border: rgba(43, 86, 136, .13);
    --slider-red: #cf4f57;
    --warm-accent: #fff5cc;
    --warm-accent-strong: #ffe680;
}


html { scroll-behavior: smooth; }
body, .stApp { font-family: 'Montserrat', sans-serif; }
.stApp { background: var(--main-surface); color: var(--ink); }
.stApp p, .stApp h1, .stApp h2, .stApp h3, .stApp h4,
.stApp label, .stApp button, .stApp input, .stApp textarea,
.stApp [data-testid="stMarkdownContainer"] { font-family: 'Montserrat', sans-serif !important; }
header[data-testid="stHeader"] { background: transparent; }
#MainMenu, footer { visibility: hidden; }
.block-container { max-width: 1460px; padding-top: 2.2rem; padding-bottom: 4rem; }

[data-testid="stSidebar"] {
    background: var(--sidebar);
    border-right: 1px solid rgba(43, 86, 136, .10);
}
[data-testid="stSidebar"] > div:first-child { padding-top: 2.8rem; }
[data-testid="stSidebarNav"] { display: none; }
[data-testid="stSidebar"] [data-testid="stIconMaterial"],
[data-testid="stSidebar"] span[class*="material-symbols"] {
    font-family: 'Material Symbols Rounded', 'Material Symbols Outlined' !important;
}
.sidebar-nav a,
.sidebar-methodology a {
    display: block;
    color: #ffffff !important;
    text-decoration: none !important;
    font-size: 14px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: .03em;
    transition: opacity .15s ease, transform .15s ease;
}
.sidebar-nav a { margin: 0 0 20px 0; }
.sidebar-nav a:nth-child(4) { margin-top: 24px; }
.sidebar-nav a:hover,
.sidebar-methodology a:hover { color: #ffffff !important; opacity: .78; transform: translateX(3px); }
.sidebar-methodology { position: fixed; bottom: 64px; width: 235px; }


[data-testid="stSidebar"] .stButton > button {
    background: transparent !important; border: 0 !important; box-shadow: none !important;
    color: #ffffff !important; padding: 0 !important; min-height: 0 !important; height: auto !important;
    justify-content: flex-start !important; font-size: 14px !important; font-weight: 400 !important;
    text-transform: uppercase; letter-spacing: .03em; border-radius: 0 !important;
}
[data-testid="stSidebar"] .stButton > button:hover { color:#ffffff !important; opacity:.78; transform:translateX(3px); }
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0 !important; }
[data-testid="stSidebar"] [data-testid="stButton"] { margin-bottom: 10px; }
.st-key-sidebar_priorities_gap {
    height: 30px;
    margin-top: 50px;
    border-top: 5px solid #FFF5CCAA;
    padding: 10px;
    width: 50%;
}
.st-key-sidebar_methodology_button { position:fixed; bottom:64px; width:235px; }

.st-key-top_nav [data-testid="stHorizontalBlock"] { align-items:flex-start; }
.st-key-top_nav [data-testid="column"] { min-width:0 !important; }
.st-key-top_nav .stButton > button {
    position:relative; background:transparent !important; border:0 !important; box-shadow:none !important;
    border-radius:0 !important; padding:0 0 15px 0 !important; min-height:0 !important; height:auto !important;
    color:var(--ink-muted) !important; font-size:12px !important; font-weight:600 !important;
    letter-spacing:.04em; text-transform:uppercase; white-space:nowrap;
}
.st-key-top_nav .stButton > button:hover { color:var(--accent) !important; box-shadow:inset 0 -5px 0 var(--warm-accent) !important; }
[data-testid="stSidebar"] .stButton > button * { font-weight: 500 !important; }
.st-key-top_nav .stButton > button * { font-weight: 700 !important; }
.st-key-nav_overall_active .stButton > button,
.st-key-nav_market_active .stButton > button,
.st-key-nav_priorities_active .stButton > button {
    color:var(--accent) !important; box-shadow:inset 0 -5px 0 var(--warm-accent) !important;
}
.st-key-period_slot { height:55px; padding-top:10px; }
.st-key-period_slot [data-testid="stHorizontalBlock"] { align-items:flex-start; }
.st-key-period_slot .stButton > button {
    border:1px solid rgba(43,86,136,.25) !important; border-radius:999px !important;
    background:rgba(255,255,255,.48) !important; color:var(--accent) !important;
    padding:7px 13px !important; min-height:0 !important; height:auto !important;
    font-size:12px !important; font-weight:600 !important; white-space:nowrap;
}
.st-key-period_slot .stButton > button:hover,
.st-key-period_q1_active .stButton > button,
.st-key-period_q2_active .stButton > button { background:var(--accent-pale) !important; }
.st-key-period_clear .stButton > button { border:0 !important; background:transparent !important; text-decoration:underline !important; }

.main-title-row { display:flex; align-items:baseline; gap:8px; position:relative; width:100%; }
.main-title {
    color: var(--ink);
    font-weight: 600;
    font-size: clamp(31px, 2.45vw, 39px);
    line-height: 1.12;
    letter-spacing: -.025em;
    margin: 0;
}
.main-subtitle { font-size: 17px; color: var(--ink-muted); margin: 10px 0 30px 0; }
.info-details { display:inline-block; position:relative; }
.info-details summary {
    list-style:none; cursor:pointer; color:var(--accent-mid); font-size:19px; line-height:1;
    user-select:none; outline:none;
}
.info-details summary::-webkit-details-marker { display:none; }
.info-details summary:hover { color:var(--accent); }
.info-panel {
    position:absolute; right:0; top:30px; z-index:1000; width:min(520px, 82vw);
    padding:18px 20px; background:#ffffff; border:1px solid rgba(43,86,136,.18);
    border-radius:12px; box-shadow:0 18px 50px rgba(43,86,136,.18);
    font-size:15px !important; line-height:1.58; color:var(--ink);
}
.info-panel p { margin:0 0 10px 0; font-size:15px !important; }
.info-panel a { color:var(--ink) !important; font-weight:700; text-decoration:underline !important; text-decoration-color:var(--warm-accent-strong) !important; text-decoration-thickness:2px !important; text-underline-offset:3px; }
.info-panel p:last-child { margin-bottom:0; }

.top-nav {
    display: flex; flex-wrap: wrap; align-items: flex-start; gap: 42px; margin: 0 0 4px 0;
}
.top-nav a {
    position: relative; display: inline-block; padding-bottom: 15px; color: var(--ink-muted) !important;
    text-decoration: none !important; font-size: 12px; font-weight: 600; letter-spacing: .04em;
    text-transform: uppercase; cursor: pointer;
}
.top-nav a::after {
    content: ''; position: absolute; left: 0; bottom: 0; width: 100%; height: 5px; border-radius: 99px;
    background: var(--warm-accent); transform: scaleX(0); transform-origin: left; transition: transform .16s ease;
}
.top-nav a:hover, .top-nav a.active { color: var(--accent) !important; }
.top-nav a:hover::after, .top-nav a.active::after { transform: scaleX(1); }
.period-slot { height: 55px; display: flex; align-items: flex-start; padding-top:10px; }
.period-nav { display: flex; gap: 10px; margin-left: 120px; align-items: center; }
.period-nav a {
    display: inline-block; padding: 7px 13px; border: 1px solid rgba(43, 86, 136, .25);
    border-radius: 999px; color: var(--accent) !important; text-decoration: none !important;
    font-size: 12px; font-weight: 600; background: rgba(255, 255, 255, .48);
}
.period-nav a:hover, .period-nav a.active { background: var(--accent-pale); }
.period-nav a.clear { border: 0; background: transparent; text-decoration: underline !important; }
.period-placeholder { visibility: hidden; height: 55px; }

.section-title { color: var(--ink); font-weight: 600; font-size: 26px; line-height: 1.25; margin: 0 0 12px 0; }
.section-kicker { color: var(--ink-muted); font-size: 11px; font-weight: 600; letter-spacing: .08em; text-transform: uppercase; margin: -2px 0 12px 0; }
.section-gap { height: 92px; }
.anchor { display: block; position: relative; top: -18px; visibility: hidden; }
.st-key-chart_header_slot { height: auto; min-height: 0; padding-top: 12px; padding-bottom: 0; overflow: visible; }
.st-key-chart_header_slot .section-title { margin-bottom: 5px; }
.st-key-chart_header_slot .base-chart-title { margin-bottom: 10px !important; }
.st-key-chart_header_slot [data-testid="stToggle"] { margin-top: -2px !important; margin-bottom: -18px !important; }
.st-key-ranking_chart_slot { margin-top: -4px !important; }
.st-key-ranking_chart_slot [data-testid="stPlotlyChart"] { margin-top: 0 !important; padding-top: 0 !important; }

.st-key-ranking_section, .st-key-table_section, .st-key-assessment_section, .st-key-priorities_section, .st-key-methodology_section, .st-key-market_drivers_section, .st-key-market_developments_section {
    background: var(--card); border: 1px solid var(--card-border); border-radius: 16px; padding: 28px;
    box-shadow: 0 8px 24px rgba(43, 86, 136, .05);
}
.st-key-assessment_section { padding-bottom: 50px; }
.st-key-priorities_section [data-testid="stForm"] { padding: 50px !important; }
.st-key-priorities_section [data-testid="stFormSubmitButton"] { margin-top: 34px; }

.stButton > button, .stFormSubmitButton > button, .stDownloadButton > button {
    border-radius: 8px !important; border-color: rgba(43, 86, 136, .30) !important; font-weight: 600 !important;
}
.stButton > button:hover, .stFormSubmitButton > button:hover, .stDownloadButton > button:hover {
    border-color: var(--accent) !important; color: var(--accent) !important;
}
[data-testid="stToggle"] label { color: var(--ink) !important; }
[data-testid="stToggle"]:hover div[role="checkbox"] { background: var(--accent-pale) !important; }
[data-testid="stToggle"] div[role="checkbox"][aria-checked="true"] { background: var(--accent) !important; }

[data-testid="stElementToolbar"] { opacity: 1 !important; background: transparent !important; }
[data-testid="stElementToolbar"] button { color: var(--accent) !important; background: transparent !important; opacity: 1 !important; }
[data-testid="stElementToolbar"] button:hover { background: var(--warm-accent) !important; }
.modebar-container, .modebar { opacity: 1 !important; background: transparent !important; }
.modebar-group { background: transparent !important; }
.modebar-btn { background: transparent !important; border-radius: 6px !important; }
.modebar-btn:hover { background: var(--warm-accent) !important; }
.modebar-btn svg { width: 19px !important; height: 19px !important; fill: var(--accent) !important; }
.modebar-btn:hover svg { fill: var(--accent) !important; }

.priority-form-space { height: 38px; }
.priority-label { color: var(--ink); font-size: 18px; font-weight: 600; margin: 18px 0 3px 0; }
.priority-description { color: var(--ink-muted); font-size: 13px; line-height: 1.5; margin: 0 0 4px 0; max-width: 900px; }
.priority-scale-note {
    display: flex; justify-content: space-between; color: var(--ink-muted); font-size: 10px; font-weight: 600;
    letter-spacing: .07em; text-transform: uppercase; margin: 10px 0 2px 0;
}
[data-testid="stSlider"] { margin-bottom: 1rem; }
[data-testid="stSlider"] [data-baseweb="slider"] > div > div {
    min-height: 11px !important; border-radius: 99px !important; background: var(--slider-red) !important;
}
[data-testid="stSlider"] [role="slider"] {
    width: 27px !important; height: 27px !important; background: var(--slider-red) !important;
    border: 3px solid #ffffff !important; box-shadow: 0 0 0 2px rgba(207, 79, 87, .25) !important;
}

.metric-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 10px 0 34px 0; }
.metric-card { background: rgba(121, 166, 210, .16); border: 1px solid rgba(43, 86, 136, .13); border-radius: 10px; padding: 14px 16px; }
.metric-label { color: var(--ink-muted); font-size: 12px; font-weight: 600; margin-bottom: 5px; }
.metric-value { color: var(--accent); font-size: 28px; font-weight: 600; line-height: 1.1; }
.assessment-heading, .priority-heading { color: var(--ink); font-size: 17px; font-weight: 600; margin: 0 0 4px 0; }
.assessment-copy { color: var(--ink); font-size: 14px; line-height: 1.68; }
.assessment-copy strong { color: var(--accent); font-weight: 700; }
.priority-heading { margin-top: 30px; margin-bottom: 12px; }
.priority-card {
    padding: 16px 18px; border-left: 5px solid var(--accent-mid); background: rgba(121, 166, 210, .13);
    border-radius: 0 9px 9px 0; color: var(--ink); font-size: 14px; line-height: 1.55; margin: 0;
}
.priority-card p { margin: 0; }
.priority-card p + p { margin-top: 12px; }
.st-key-country_picker { width: 100%; max-width: 100%; }
.st-key-country_picker div[data-baseweb="select"],
.st-key-industry_picker div[data-baseweb="select"],
.st-key-role_picker div[data-baseweb="select"] { font-size: 16px; }
.st-key-country_picker div[data-baseweb="select"] > div,
.st-key-industry_picker div[data-baseweb="select"] > div,
.st-key-role_picker div[data-baseweb="select"] > div { background: var(--warm-accent) !important; }
.st-key-assessment_bottom_row [data-testid="stHorizontalBlock"] { align-items: stretch; }
.st-key-assessment_bottom_row [data-testid="column"] { display: flex; flex-direction: column; }
.st-key-priority_panel, .st-key-pillar_explanation_panel { height: 100%; display: flex; flex-direction: column; justify-content: flex-end; }
.pillar-legend-box {
    margin-top: 0; padding: 18px 20px; border: 1px solid var(--warm-accent-strong); border-radius: 10px;
    display: grid; grid-template-columns: 1fr; gap: 15px; font-size: 12px; line-height: 1.4; color: var(--ink);
}
.pillar-definition strong { display: block; margin-bottom: 5px; }
.pillar-definition p { margin: 0; color: var(--ink-muted); line-height: 1.5; }
.assessment-footnote { margin-top:28px; padding-top:14px; border-top:1px solid rgba(43,86,136,.12); color:var(--ink-muted); font-size:11px; line-height:1.45; }

.chart-explainer {
    margin: 25px auto 25px auto;
    padding: 0;
    max-width: 70%;
    border: 0;
    background: transparent;
    color: var(--ink);
    font-size: 14px;
    line-height: 1.68;
    text-align: left;
}
.chart-explainer p { margin: 0; }
.chart-explainer p + p { margin-top: 7px; }
.chart-explainer strong { color: var(--ink); font-weight: 600; }
.market-intro {
    width:100%;
    max-width:none;
    margin:0 0 24px 0;
    color:var(--ink);
    font-size:14px;
    line-height:1.68;
    text-align:left;
}
.market-driver-table {
    width:100%;
    table-layout:fixed;
    border-collapse:collapse;
    margin-top:10px;
    font-size:13px;
    line-height:1.5;
}
.market-driver-table th,
.market-driver-table td {
    overflow-wrap:break-word;
    word-break:normal;
}
.market-driver-table th {
    padding:10px 12px;
    text-align:left;
    vertical-align:middle;
    background:#fffae6;
    font-weight:600;
    border:1px solid rgba(43,86,136,.12);
}
.market-driver-table td {
    padding:12px;
    vertical-align:top;
    border:1px solid rgba(43,86,136,.10);
}
.market-confidence-effect { color:var(--accent); font-weight:600; }
.market-context-copy {
    margin-top: 24px;
    padding-bottom: 25px !important;
    color: var(--ink);
    font-size: 15px !important;
    line-height: 1.55;
}

.market-context-copy .market-context-heading {
    margin: 0 0 12px 0;
    color: var(--ink);
    font-size: 15px !important;
    font-weight: 500;
    line-height: 1.4;
}

.market-context-copy p {
    margin: 0 0 12px 0;
    font-size: 15px !important;
    line-height: 1.55;
}

.market-context-copy .market-context-list {
    margin: 15px 0 0 20px;
    padding: 0;
    font-size: 15px !important;
    line-height: 1.55;
}

.market-context-copy .market-context-list li {
    margin-bottom: 10px;
    font-size: 15px !important;
}

.market-context-copy .market-context-list li:last-child {
    margin-bottom: 0;
}

.market-context-copy strong {
    color: var(--accent);
    font-weight: 700;
}
.market-development-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px; margin-top:14px; }
.market-development-card { padding:18px 20px; border:1px solid rgba(43,86,136,.14); border-radius:11px; background:rgba(255,255,255,.82); }
.market-development-title { font-size:15px; font-weight:600; line-height:1.45; color:var(--ink); margin-bottom:6px; }
.market-development-meta { color:var(--ink-muted); font-size:11px; line-height:1.45; margin-bottom:10px; }
.market-development-meta a {
    color:var(--accent) !important;
    font-weight:600;
    text-decoration:underline !important;
    text-decoration-color:var(--warm-accent-strong) !important;
    text-decoration-thickness:2px !important;
    text-underline-offset:3px;
}
.market-development-meta a:hover { opacity:.76; }
.market-development-summary { color:var(--ink); font-size:13px; line-height:1.58; }
.market-development-map { margin-top:10px; color:var(--ink-muted); font-size:11px; line-height:1.45; }
.st-key-market_main_link { margin-top:24px; text-align:center; }
.st-key-market_main_link .stButton > button { background:transparent !important; border:0 !important; text-decoration:underline !important; color:var(--accent) !important; }
.market-development-signal {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-top: 18px;
    color: var(--accent);
    font-size: 18px;
    font-weight: 700;
    line-height: 1.3;
}

.market-development-signal-icon {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 28px;
    height: 28px;
    flex: 0 0 28px;
    border: 1.5px solid currentColor;
    border-radius: 50%;
    font-size: 18px;
    font-weight: 700;
    line-height: 1;
}

@media (max-width: 768px) {
    .market-development-signal {
        font-size: 16px;
    }

    .market-development-signal-icon {
        width: 26px;
        height: 26px;
        flex-basis: 26px;
        font-size: 16px;
    }
}

.validation-shell {
    max-width: 1020px;
    color: var(--ink);
    font-size: 14px;
    line-height: 1.65;
}
.validation-access-card,
.validation-consent-card,
.validation-question-card {
    padding: 18px 20px;
    border: 1px solid rgba(43,86,136,.14);
    border-radius: 12px;
    background: rgba(255,255,255,.84);
    margin: 14px 0;
}
.validation-code {
    display: inline-block;
    padding: 5px 9px;
    border-radius: 6px;
    background: var(--warm-accent);
    color: var(--accent);
    font-size: 12px;
    font-weight: 700;
    letter-spacing: .04em;
}
.validation-note {
    color: var(--ink-muted);
    font-size: 12px;
    line-height: 1.55;
}
.validation-question-label {
    color: var(--ink);
    font-size: 15px;
    font-weight: 600;
    margin-bottom: 5px;
}
.validation-question-description {
    color: var(--ink-muted);
    font-size: 12px;
    line-height: 1.5;
    margin-bottom: 8px;
}

.action-progress-wrap {
    min-height: 46px;
    margin: 18px 0 10px 0;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 12px;
    color: var(--ink-muted);
    font-size: 13px;
    font-weight: 500;
}
.action-progress-spinner {
    width: 24px;
    height: 24px;
    border: 2px solid rgba(43, 86, 136, .16);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: priority-spin .75s linear infinite;
}
.st-key-custom_actions {
    position: sticky;
    bottom: 12px;
    z-index: 200;
    margin-top: 6px;
    padding: 18px 0 6px 0;
    background: rgba(255,255,255,.98);
    border: 0;
    border-top: 1px solid var(--warm-accent-strong);
    box-shadow: none;
    border-radius: 0;
}
.st-key-custom_actions .stButton > button,
.st-key-custom_actions .stDownloadButton > button { width: 100% !important; min-width: 132px !important; }
.action-warning {
    max-width: 620px; margin: 0 auto 8px auto; padding: 9px 12px; border-radius: 7px;
    background: #fffae6; color: var(--ink); font-size: 12px; line-height: 1.45; text-align: center;
}
.action-message {
    max-width: 760px;
    margin: 20px auto;
    padding: 0;
    border: 0 !important;
    border-radius: 0;
    background: transparent !important;
    color: var(--accent) !important;
    font-size: 12px;
    font-weight: 600;
    line-height: 1.45;
    text-align: center;
}
.action-message.success,
.action-message.error,
.action-message.warning {
    border: 0 !important;
    background: transparent !important;
    color: var(--accent) !important;
}
.saved-line { color: var(--ink-muted); font-size: 12px; margin: 12px 0 0 0; text-align: center; }
.form-note, .backend-note { color: var(--ink-muted); font-size: 12px; }
.backend-note { padding: 10px 12px; background: rgba(121, 166, 210, .14); border-radius: 7px; }
.gate-copy { font-size:14px; line-height:1.55; color:var(--ink); margin-bottom:10px; }
.gate-fine-print { font-size:10px; line-height:1.45; color:var(--ink-muted); margin-top:10px; margin-bottom:24px; }
.st-key-gate_primary_button .stButton > button,
.st-key-gate_primary_button .stFormSubmitButton > button { width:100% !important; }
.st-key-gate_decline_button .stButton > button { border:0 !important; background:transparent !important; box-shadow:none !important; font-size:13px !important; font-weight:600 !important; text-decoration:underline !important; color:var(--ink) !important; }
.gate-inline-progress { min-height:90px; display:flex; flex-direction:column; align-items:center; justify-content:center; gap:12px; color:var(--accent); font-size:13px; font-weight:600; }
.gate-progress-wrap {
    min-height: 180px;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 14px;
    color: var(--accent);
    font-size: 13px;
    font-weight: 600;
    text-align: center;
}
.gate-progress-spinner {
    width: 28px;
    height: 28px;
    border: 2px solid rgba(43, 86, 136, .16);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: priority-spin .75s linear infinite;
}
[data-testid="stDialog"], [data-testid="stDialog"] * { font-family:'Montserrat', sans-serif !important; }
[data-testid="stDialog"] [data-testid="stFormSubmitButton"] { margin-top: 18px; }
.priority-processing {
    min-height: 260px; display: flex; flex-direction: column; align-items: center; justify-content: center;
    color: var(--ink-muted); font-size: 13px; gap: 14px;
}
.priority-processing-spinner {
    width: 28px; height: 28px; border: 2px solid rgba(43,86,136,.18); border-top-color: var(--accent);
    border-radius: 50%; animation: priority-spin .8s linear infinite;
}
@keyframes priority-spin { to { transform: rotate(360deg); } }
.methodology-copy { max-width: 1020px; font-size: 15px; line-height: 1.72; color: var(--ink); }
.methodology-copy p { margin: 0 0 16px 0; }
.methodology-copy ul { margin: 0 0 18px 22px; padding: 0; }
.methodology-copy li { margin: 0 0 6px 0; }
.methodology-subhead { font-size: 19px; font-weight: 500; line-height: 1.35; margin: 30px 0 12px 0; color: var(--ink); }

@media (max-width: 900px) {
    .main-title { font-size: 31px; }
    .main-subtitle { font-size: 16px; }
    .top-nav { gap: 24px; }
    .period-nav { margin-left: 0; }
    .sidebar-methodology { position: static; width: auto; margin-top: 46px; }
    .st-key-ranking_section, .st-key-table_section, .st-key-assessment_section, .st-key-priorities_section, .st-key-methodology_section, .st-key-market_drivers_section, .st-key-market_developments_section { padding: 20px 18px; }
    .pillar-legend-box { grid-template-columns:1fr; }
}

/* Expanded methodology page */
.methodology-copy .pillar-heading { font-size:15px; font-weight:700; line-height:1.5; margin:28px 0 8px; }
.methodology-copy .indicator-list { margin-top:8px !important; }
.methodology-copy .indicator-list li { margin-bottom:12px; line-height:1.6; }
.methodology-copy a { color:var(--accent) !important; font-weight:600; text-decoration:underline !important; text-decoration-color:var(--warm-accent-strong) !important; text-decoration-thickness:2px !important; text-underline-offset:3px; }
.methodology-note { margin:18px 0; padding:14px 16px; border-left:4px solid var(--accent-mid); background:rgba(121,166,210,.12); font-size:13px; line-height:1.55; }
.methodology-weight-table { width:100%; border-collapse:collapse; margin:12px 0 20px; font-size:13px; }
.methodology-weight-table th, .methodology-weight-table td { padding:9px 10px; text-align:left; vertical-align:top; border-bottom:1px solid rgba(43,86,136,.12); }
.methodology-weight-table th { font-weight:700; background:#fffae6; }
.validation-box { width:70%; margin:20px auto 8px auto; padding:18px 20px; border:1px solid var(--warm-accent-strong); border-radius:10px; background:rgba(255,255,255,.78); }
.validation-title { font-size:17px; font-weight:500; margin-bottom:10px; }
.validation-grid { display:grid; grid-template-columns:minmax(260px,1fr) auto; gap:8px 24px; font-size:13px; line-height:1.45; }
.validation-value { color:var(--accent); font-weight:700; text-align:right; }
.methodology-footer { margin-top:30px !important; padding-top:14px; border-top:1px solid rgba(43,86,136,.12); color:var(--ink-muted); font-size:10px; line-height:1.45; letter-spacing:.015em; text-transform:uppercase; }
@media (max-width:760px) { .validation-grid { grid-template-columns:1fr; } .validation-value { text-align:left; margin-bottom:7px; } .validation-box { width:100%; } .methodology-footer { font-size:10px; } .market-development-grid { grid-template-columns:1fr; } }



/* Research survey */
.st-key-validation_section {
    background: var(--card);
    border: 1px solid var(--card-border);
    border-radius: 16px;
    padding: 28px;
    box-shadow: 0 8px 24px rgba(43, 86, 136, .05);
}
.st-key-validation_form_shell { padding: 18px 50px 50px 50px; }
.survey-intro { max-width: 1000px; font-size: 16px; line-height: 1.68; color: var(--ink); }
.survey-intro p { margin: 0 0 13px 0; }
.survey-time { color: var(--ink-muted); font-size: 12px; font-weight: 600; margin-top: 10px; }
.survey-required-note { color: var(--ink-muted); font-size: 11px; line-height: 1.5; margin: 10px 0 0 0; }
.survey-subtitle { color: var(--ink); font-size: 19px; font-weight: 600; line-height: 1.35; margin: 28px 0 12px 0; }
.survey-help { color: var(--ink-muted); font-size: 12px; line-height: 1.55; margin: 4px 0 14px 0; }
.survey-privacy { color: var(--ink-muted); font-size: 12px; line-height: 1.6; }
.survey-privacy-before-consent { margin: 12px 0 3px 0; }
.survey-privacy-after-consent { margin: 1px 0 30px 0; }
.st-key-validation_contact_consent_row { margin: 0 !important; padding: 0 !important; }
.st-key-validation_contact_consent_row [data-testid="stCheckbox"] { margin: 0 !important; padding: 0 !important; }
.survey-details-gap { height: 50px; }
.priority-dropdown-gap { height: 35px; }
.survey-question-intro { font-size: 17px; font-weight: 600; line-height: 1.45; margin: 0 0 10px 0; color: var(--ink); }
.survey-scenario-copy { color: var(--ink-muted); font-size: 15px; line-height: 1.68; margin: 0 0 22px 0; }
.survey-scenario-copy p { margin: 0 0 8px 0; font-size: 15px !important; line-height: 1.68 !important; }
.survey-scenario-copy p:last-child { margin-bottom: 0; }
.survey-later-question { margin-top: 38px; }
.survey-counter { color: var(--accent); font-size: 12px; font-weight: 700; margin: -4px 0 12px 0; }
.survey-footnote { color: var(--ink-muted); font-size: 11px; line-height: 1.5; margin-top: 8px; }
.survey-optional-label { color: var(--ink-muted); font-size: 11px; font-weight: 500; margin-left: 6px; }
.st-key-survey_submit_row { margin-top: 34px; }
.st-key-survey_submit_row .stButton > button { min-width: 132px; width: auto !important; }

/* Survey information popovers. Native popovers close on outside tap and Escape. */
.survey-info-trigger {
    appearance: none;
    -webkit-appearance: none;
    border: 0;
    padding: 0;
    background: transparent;
    font-family: 'Montserrat', sans-serif !important;
    cursor: pointer;
}
.survey-info-link-trigger {
    display: block;
    width: fit-content;
    margin: 12px 0 2px 0;
    color: var(--accent) !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    line-height: 1.45 !important;
    text-decoration: underline;
    text-decoration-color: var(--warm-accent-strong);
    text-decoration-thickness: 3px;
    text-underline-offset: 3px;
}
.survey-info-icon-trigger {
    display: inline-block !important;
    margin-left: 5px;
    color: var(--accent-mid) !important;
    font-size: 16px !important;
    line-height: 1 !important;
    vertical-align: baseline;
}
.survey-info-popover {
    position: fixed !important;
    inset: 0 !important;
    width: min(680px, calc(100vw - 36px)) !important;
    max-width: 680px !important;
    max-height: min(72vh, 620px) !important;
    margin: auto !important;
    padding: 58px 24px 22px 24px !important;
    overflow-y: auto !important;
    border: 1px solid rgba(43,86,136,.18) !important;
    border-radius: 14px !important;
    background: #ffffff !important;
    box-shadow: 0 20px 60px rgba(25,45,70,.24) !important;
    color: var(--ink) !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 15px !important;
    line-height: 1.58 !important;
    z-index: 5000 !important;
}
.survey-info-popover::backdrop {
    background: rgba(19, 35, 54, .22);
}
.survey-info-popover p,
.survey-info-popover a,
.survey-info-popover strong {
    font-family: 'Montserrat', sans-serif !important;
}
.survey-info-popover p {
    margin: 0 0 12px 0 !important;
    font-size: 15px !important;
    line-height: 1.58 !important;
}
.survey-info-popover p:last-child { margin-bottom: 0 !important; }
.survey-info-close {
    position: absolute;
    top: 12px;
    right: 14px;
    min-width: 92px;
    min-height: 36px;
    padding: 7px 12px;
    border: 1px solid rgba(43,86,136,.30);
    border-radius: 8px;
    background: var(--accent) !important;
    color: #ffffff !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 13px;
    font-weight: 700;
    line-height: 1;
    cursor: pointer;
}
.survey-info-close:hover,
.survey-info-close:focus-visible {
    background: #1f4775 !important;
    outline: 3px solid var(--warm-accent-strong);
    outline-offset: 2px;
}
.survey-info-close-x {
    margin-left: 6px;
    font-size: 17px;
    line-height: .8;
    vertical-align: -1px;
}
.survey-scenario-lead { display: block; }
.survey-scenario-lead > span { display: inline; }

/* More breathing room between radio options. */
.st-key-validation_section [data-testid="stRadio"] [role="radiogroup"] { gap: 9px !important; }
.st-key-validation_section [data-testid="stRadio"] label { line-height: 1.55 !important; margin-bottom: 3px !important; }

/* Research-survey invitation shown only after a customised priorities result. */
.chart-explainer .research-survey-invite { margin-top: 13px; font-size: 14px; line-height: 1.55; }
.chart-explainer .research-survey-invite a {
    color: var(--accent) !important;
    font-size: 14px;
    font-weight: 600;
    text-decoration: underline !important;
    text-decoration-color: var(--warm-accent-strong) !important;
    text-decoration-thickness: 3px !important;
    text-underline-offset: 3px;
}

/* A direct survey link is intentionally distraction-free. */
.survey-direct-mode [data-testid="stSidebar"] { display: none !important; }


/* Survey importance scales: one full-width scale, with both labels aligned above the bar. */
.st-key-validation_section .priority-description {
    margin-bottom: 20px !important;
}
[class*="st-key-validation_factor_scale_"] {
    margin: 0 0 2.35rem 0 !important;
}
[class*="st-key-validation_factor_scale_"] [data-testid="stSlider"] {
    width: 100% !important;
    min-width: 0 !important;
    max-width: none !important;
    margin: 0 !important;
}
.survey-scale-label-row {
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    width: 100%;
    margin: 0 0 5px 0;
    color: var(--ink-muted);
    font-size: 10px;
    font-weight: 600;
    line-height: 1.25;
    letter-spacing: .055em;
    text-transform: uppercase;
}
.survey-scale-label-row span:last-child { text-align: right; }
/* Streamlit also renders its own lower endpoint labels. The script below hides those only. */
.survey-hide-slider-label {
    display: none !important;
}

/* BaseWeb mounts some select and multiselect text outside the normal app tree. */
.st-key-validation_section [data-testid="stMultiSelect"],
.st-key-validation_section [data-testid="stMultiSelect"] *,
.st-key-validation_section [data-testid="stSelectbox"],
.st-key-validation_section [data-testid="stSelectbox"] *,
.st-key-validation_section [data-baseweb="select"],
.st-key-validation_section [data-baseweb="select"] *,
div[data-baseweb="popover"],
div[data-baseweb="popover"] *,
ul[role="listbox"],
ul[role="listbox"] * {
    font-family: 'Montserrat', sans-serif !important;
}

/* Two-column overview for the five-factor selection. */
.st-key-validation_top_five_grid [data-testid="stHorizontalBlock"] {
    align-items: flex-start !important;
}
.st-key-validation_top_five_grid [data-testid="stCheckbox"] {
    margin: 0 0 8px 0 !important;
}
.st-key-validation_top_five_grid [data-testid="stCheckbox"] label {
    line-height: 1.45 !important;
}

/* Mobile notice for the analytical app. The research survey remains mobile-friendly. */
.mobile-desktop-notice {
    display: none;
}

/* Compact, readable validation summary. */
.st-key-validation_section [data-testid="stAlert"] {
    font-family: 'Montserrat', sans-serif !important;
    font-size: 13px !important;
    line-height: 1.55 !important;
}
.st-key-validation_section [data-testid="stAlert"] ul {
    margin: 8px 0 0 18px !important;
    padding: 0 !important;
}
.st-key-validation_section [data-testid="stAlert"] li { margin: 0 0 5px 0 !important; }

@media (max-width: 900px) {
    .mobile-desktop-notice {
        display: block !important;
        margin: 0 0 18px 0;
        padding: 16px 18px;
        border: 1px solid var(--warm-accent-strong);
        border-radius: 10px;
        background: #fffae6;
        color: var(--ink);
        font-family: 'Montserrat', sans-serif !important;
        font-size: 15px;
        line-height: 1.55;
    }
    .mobile-desktop-notice strong {
        display: block;
        margin-bottom: 5px;
        color: var(--accent);
        font-weight: 700;
    }
    .block-container {
        max-width: 100% !important;
        padding: 1rem .75rem 2.5rem .75rem !important;
    }
    .st-key-validation_section {
        padding: 20px 16px !important;
        border-radius: 12px !important;
    }
    .st-key-validation_form_shell { padding: 8px 0 30px 0 !important; }
    .st-key-validation_section .section-title {
        font-size: 25px !important;
        line-height: 1.2 !important;
    }
    .survey-intro { font-size: 14px !important; line-height: 1.6 !important; }
    .survey-subtitle { font-size: 17px !important; margin-top: 26px !important; }
    .survey-question-intro { font-size: 16px !important; }
    .survey-scenario-copy,
    .survey-scenario-copy p { font-size: 14px !important; line-height: 1.62 !important; }
    .survey-details-gap { height: 34px !important; }
    .priority-label { font-size: 16px !important; line-height: 1.35 !important; }
    .priority-description { font-size: 13px !important; line-height: 1.5 !important; }

    /* Stack all two- and three-column survey rows on smaller screens. */
    .st-key-validation_form_shell [data-testid="stHorizontalBlock"] {
        flex-direction: column !important;
        gap: .7rem !important;
    }
    .st-key-validation_form_shell [data-testid="column"] {
        width: 100% !important;
        flex: 1 1 100% !important;
        min-width: 0 !important;
    }

    .st-key-validation_section input,
    .st-key-validation_section textarea,
    .st-key-validation_section [data-baseweb="select"] {
        font-size: 16px !important;
    }
    .st-key-validation_section [data-testid="stRadio"] [role="radiogroup"] {
        gap: 12px !important;
    }
    .st-key-validation_section [data-testid="stRadio"] label {
        align-items: flex-start !important;
        font-size: 14px !important;
        line-height: 1.6 !important;
        margin-bottom: 5px !important;
    }
    /* Full-width importance scales remain aligned on mobile. */
    [class*="st-key-validation_factor_scale_"] [data-testid="stSlider"] {
        width: 100% !important;
        min-width: 0 !important;
        max-width: none !important;
        margin: 0 !important;
    }
    .survey-scale-label-row {
        font-size: 8px !important;
        letter-spacing: .025em !important;
        line-height: 1.2 !important;
        margin-bottom: 6px !important;
        white-space: nowrap !important;
    }
    .st-key-validation_section [data-testid="stMultiSelect"] { width: 100% !important; }

    /* Compact mobile popover: visible close control, smaller type and spare screen space. */
    .survey-info-popover {
        width: calc(100vw - 28px) !important;
        max-width: none !important;
        max-height: min(68dvh, 520px) !important;
        padding: 54px 16px 18px 16px !important;
        border-radius: 12px !important;
        font-size: 13px !important;
        line-height: 1.52 !important;
        overscroll-behavior: contain;
    }
    .survey-info-popover p {
        font-size: 13px !important;
        line-height: 1.52 !important;
        margin-bottom: 10px !important;
    }
    .survey-info-close {
        top: 10px;
        right: 10px;
        min-width: 94px;
        min-height: 38px;
        font-size: 13px;
    }
    .st-key-survey_submit_row .stButton > button {
        width: 100% !important;
        min-height: 44px !important;
    }
}

@media (max-width: 480px) {
    .block-container { padding-left: .5rem !important; padding-right: .5rem !important; }
    .st-key-validation_section { padding: 17px 12px !important; }
    .st-key-validation_section .section-title { font-size: 22px !important; }
    .survey-info-link-trigger { font-size: 13px !important; }
    .priority-scale-note { font-size: 9px !important; }
}

</style>
"""

st.markdown(CSS, unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def load_index_data(file_path: str) -> pd.DataFrame:
    """Read cached Excel values without changing the workbook."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"The workbook does not contain a '{SHEET_NAME}' sheet.")

    worksheet = workbook[SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        raise ValueError("The Streamlit sheet is empty.")

    headers = list(rows[0])
    while headers and headers[-1] is None:
        headers.pop()

    data_rows: list[list[Any]] = []
    for row in rows[1:]:
        trimmed = list(row[: len(headers)])
        if any(value not in (None, "") for value in trimmed):
            data_rows.append(trimmed)

    frame = pd.DataFrame(data_rows, columns=headers)
    frame = frame[frame["Country"].notna() & frame["ISO3"].notna()].copy()

    missing = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
    if missing:
        workbook.close()
        raise ValueError("Missing required columns: " + ", ".join(missing))

    numeric_columns = ["Base index rank", "Base index score", *PILLAR_SOURCES]
    for column in numeric_columns:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")

    if frame["Base index score"].isna().any():
        workbook.close()
        raise ValueError(
            "Some formula results are unavailable. Open the workbook in Excel, allow recalculation, save it, and replace data/Index_Model.xlsx."
        )

    frame = frame[frame["Eligible"].astype(str).str.strip().str.lower().isin({"yes", "true", "1"})]

    if "Screening" in workbook.sheetnames:
        screening = workbook["Screening"]
        screening_rows = list(screening.iter_rows(values_only=True))
        if screening_rows:
            screening_headers = list(screening_rows[0])
            screening_data = pd.DataFrame(screening_rows[1:], columns=screening_headers)
            if "ISO3" in screening_data.columns and "Operating coal capacity (MW)" in screening_data.columns:
                screening_data = screening_data[["ISO3", "Operating coal capacity (MW)"]].copy()
                screening_data["Operating coal capacity (MW)"] = pd.to_numeric(
                    screening_data["Operating coal capacity (MW)"], errors="coerce"
                )
                frame = frame.merge(screening_data, on="ISO3", how="left")
    workbook.close()

    if "Operating coal capacity (MW)" not in frame.columns:
        frame["Operating coal capacity (MW)"] = pd.NA

    return frame.sort_values("Base index rank").reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_market_overlay(file_path: str) -> pd.DataFrame:
    """Read the final Q1 and Q2 country overlay results from the OVERLAY sheet."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if OVERLAY_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"The workbook does not contain an '{OVERLAY_SHEET_NAME}' sheet.")

    worksheet = workbook[OVERLAY_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if len(row) >= 2 and str(row[0]).strip() == "Country" and str(row[1]).strip() == "ISO3"
        ),
        None,
    )
    if header_index is None:
        workbook.close()
        raise ValueError("The OVERLAY sheet does not contain the final country-results table.")

    headers = list(rows[header_index])
    while headers and headers[-1] is None:
        headers.pop()

    data_rows: list[list[Any]] = []
    for row in rows[header_index + 1 :]:
        trimmed = list(row[: len(headers)])
        if not trimmed or trimmed[0] in (None, ""):
            break
        data_rows.append(trimmed)
    workbook.close()

    frame = pd.DataFrame(data_rows, columns=headers)
    required = [
        "Country",
        "ISO3",
        "Q1 overlay score",
        "Q1 rank",
        "Q1 rank change",
        "Q2 overlay score",
        "Q2 rank",
        "Q2 rank change",
    ]
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError("Missing OVERLAY columns: " + ", ".join(missing))

    numeric_columns = [
        "Q1 overlay score",
        "Q1 rank",
        "Q1 score change",
        "Q1 rank change",
        "Q2 overlay score",
        "Q2 rank",
        "Q2 rank change",
    ]
    for column in numeric_columns:
        if column in frame.columns:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")

    if frame[["Q1 overlay score", "Q1 rank", "Q2 overlay score", "Q2 rank"]].isna().any().any():
        raise ValueError("Some final Q1 or Q2 overlay results are unavailable in the OVERLAY sheet.")
    if frame["ISO3"].duplicated().any():
        raise ValueError("The OVERLAY country-results table contains duplicate ISO3 codes.")

    return frame.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_market_content(
    file_path: str,
    file_modified_time: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Read the public-facing market drivers, context signals and developments."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    def read_sheet(sheet_name: str) -> pd.DataFrame:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"The workbook does not contain a '{sheet_name}' sheet.")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        headers = list(rows[0])
        while headers and headers[-1] is None:
            headers.pop()
        data = [list(row[:len(headers)]) for row in rows[1:] if any(v not in (None, "") for v in row[:len(headers)])]
        return pd.DataFrame(data, columns=headers)

    try:
        drivers = read_sheet(MARKET_DRIVERS_SHEET)
        context = read_sheet(MARKET_CONTEXT_SHEET)
        developments = read_sheet(MARKET_DEVELOPMENTS_SHEET)
    finally:
        workbook.close()
    return drivers, context, developments


def market_results(index_data: pd.DataFrame, overlay_data: pd.DataFrame, period: str) -> pd.DataFrame:
    score_column = "Q1 overlay score" if period == "q1" else "Q2 overlay score"
    rank_column = "Q1 rank" if period == "q1" else "Q2 rank"
    rank_change_column = "Q1 rank change" if period == "q1" else "Q2 rank change"
    score_change_column = "Q1 score change" if period == "q1" else None

    result = base_results(index_data)
    overlay_columns = ["ISO3", score_column, rank_column, rank_change_column]
    if score_change_column and score_change_column in overlay_data.columns:
        overlay_columns.append(score_change_column)
    result = result.merge(overlay_data[overlay_columns], on="ISO3", how="left", validate="one_to_one")

    if result[[score_column, rank_column]].isna().any().any():
        missing = result.loc[result[score_column].isna() | result[rank_column].isna(), "ISO3"].tolist()
        raise ValueError("Missing market-overlay results for: " + ", ".join(missing))

    result["Overall score"] = result[score_column].astype(float)
    result["Rank"] = result[rank_column].astype(int)
    result["Rank change"] = result[rank_change_column].astype(int)
    if score_change_column and score_change_column in result.columns:
        result["Score change"] = result[score_change_column].astype(float)
    else:
        result["Score change"] = result["Overall score"] - result["Base score"]

    return result.sort_values(["Rank", "Country"]).reset_index(drop=True)


SINGAPORE_TZ = ZoneInfo("Asia/Singapore")

def singapore_now() -> str:
    return datetime.now(SINGAPORE_TZ).replace(microsecond=0).isoformat()


def initialise_state() -> None:
    defaults = {
        "session_id": str(uuid.uuid4()),
        "session_started_at": singapore_now(),
        "previews": [],
        "saved_preview_ids": [],
        "current_preview_id": None,
        "selected_country": "Chile",
        "contact": {"name": "", "organisation": "", "email": ""},
        "profile_complete": False,
        "gate_action": None,
        "gate_version_id": None,
        "session_logged": False,
        "scroll_to_top": False,
        "scroll_target": None,
        "notice": None,
        "action_notice": None,
        "action_notice_type": "warning",
        "logging_notice": None,
        "profile_industry": "Select your industry",
        "profile_role": "Select your role",
        "pending_priority_submission": None,
        "action_in_progress": None,
        "action_version_id": None,
        "cached_pdf": None,
        "cached_pdf_version_id": None,
        "profile_capture_pending": False,
        "profile_capture_reason": None,
        "gate_processing": False,
        "download_authorised_version_id": None,
        "validation_access_granted": False,
        "validation_participant_code": "VAL-" + uuid.uuid4().hex[:8].upper(),
        "validation_registered": False,
        "validation_assignment": [],
        "validation_notice": None,
        "validation_submitted": False,
        "validation_submission_message": None,
        "validation_submission_id": "VSUB-" + uuid.uuid4().hex[:12].upper(),
        "survey_success_pending": False,
        "survey_entry_locked": False,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def query_value(key: str, default: str) -> str:
    value = st.query_params.get(key, default)
    if isinstance(value, list):
        value = value[0] if value else default
    return str(value)


def set_view(view: str) -> None:
    if SURVEY_ROUTE_PARAM in st.query_params:
        del st.query_params[SURVEY_ROUTE_PARAM]
    st.query_params["view"] = view
    if view != "market" and "period" in st.query_params:
        del st.query_params["period"]


def reset_priority_survey() -> None:
    """Clear the current customised result while retaining the user's profile for this session."""
    st.session_state.previews = []
    st.session_state.saved_preview_ids = []
    st.session_state.current_preview_id = None
    st.session_state.gate_action = None
    st.session_state.gate_version_id = None
    st.session_state.action_notice = None
    st.session_state.action_notice_type = "warning"
    st.session_state.logging_notice = None
    st.session_state.action_in_progress = None
    st.session_state.action_version_id = None
    st.session_state.cached_pdf = None
    st.session_state.cached_pdf_version_id = None
    st.session_state.pending_priority_submission = None
    st.session_state.profile_capture_pending = False
    st.session_state.profile_capture_reason = None
    st.session_state.gate_processing = False
    st.session_state.download_authorised_version_id = None

    for item in SURVEY_ITEMS:
        st.session_state[f"priority_slider_{item['key']}"] = 3

    # Preserve the session's selected industry and role.
    st.session_state.industry_widget = st.session_state.profile_industry
    st.session_state.role_widget = st.session_state.profile_role

    if "result" in st.query_params:
        del st.query_params["result"]


def navigate_to(view: str, period: str | None = None, scroll_target: str = "top") -> None:
    if SURVEY_ROUTE_PARAM in st.query_params:
        del st.query_params[SURVEY_ROUTE_PARAM]
    st.query_params["view"] = view
    if period:
        st.query_params["period"] = period
    elif "period" in st.query_params:
        del st.query_params["period"]
    if "result" in st.query_params:
        del st.query_params["result"]
    st.session_state.scroll_target = scroll_target
    st.session_state.scroll_to_top = scroll_target == "top"
    st.rerun()


def render_sidebar(view: str) -> None:
    with st.sidebar:
        if st.button("Readiness Index", key="sidebar_readiness", type="tertiary"):
            navigate_to("overall", scroll_target="top")
        if st.button("Pillar Scores", key="sidebar_pillars", type="tertiary"):
            navigate_to("overall", scroll_target="pillar-scores")
        if st.button("Country Assessment", key="sidebar_assessment", type="tertiary"):
            navigate_to("overall", scroll_target="country-assessment")
        with st.container(key="sidebar_priorities_gap"):
            st.markdown("", unsafe_allow_html=True)
        if st.button("Add Your Priorities", key="sidebar_priorities", type="tertiary"):
            if view != "priorities":
                reset_priority_survey()
            navigate_to("priorities", scroll_target="top")
        with st.container(key="sidebar_methodology_button"):
            if st.button("Methodology", key="sidebar_methodology", type="tertiary"):
                navigate_to("methodology", scroll_target="top")


def render_header() -> None:
    st.markdown('<span id="top" class="anchor"></span>', unsafe_allow_html=True)
    header_html = """
    <div class="main-title-row">
      <div class="main-title">Coal-to-Clean Jurisdictional Readiness Index 2026</div>
      <details class="info-details" id="index-info-details">
        <summary aria-label="About the index" title="About the index">ⓘ</summary>
        <div class="info-panel">
          <p><strong>About the index</strong></p>
          <p>The Coal-to-Clean Jurisdictional Readiness Index compares coal-dependent markets on their readiness to support credible coal-to-clean energy transition opportunities.</p>
          <p>The index combines six pillars covering energy system conditions, policy commitment, governance capacity, carbon market maturity, macro-financial conditions, and just transition credibility. Scores are calculated from normalised indicators and weighted to produce an overall jurisdictional readiness score.</p>
          <p>The base index primarily reflects 2024 data, using the latest available year only where 2024 values were unavailable.</p>
          <p>The index is a jurisdiction-level screening tool. It does not assess the viability of individual power plants, projects or investments.</p>
          <p style="margin-top:16px;"><strong>About the market confidence overlay</strong></p>
          <p>The market-confidence overlay adds more recent evidence to the underlying index. It adjusts selected indicator weights where market developments suggest that certain readiness factors have become more or less important.</p>
          <p>The overlay does not replace the base index or alter its underlying data. It provides a separate, time-specific view for Q1 2026 and Q2 2026.</p>
          <p>Survey-based user weightings are calculated separately and are not combined with the main market-confidence overlay.</p>
          <p style="margin-top:18px;">For more information, please see <a href="?view=methodology#top" class="methodology-link">Methodology</a>.</p>
        </div>
      </details>
    </div>
    <script>
      (() => {
        const details = document.getElementById('index-info-details');
        if (!details) return;
        document.addEventListener('click', (event) => {
          if (details.open && !details.contains(event.target)) details.removeAttribute('open');
        });
        document.addEventListener('keydown', (event) => {
          if (event.key === 'Escape') details.removeAttribute('open');
        });
        const methodologyLink = details.querySelector('.methodology-link');
        if (methodologyLink) {
          methodologyLink.addEventListener('click', (event) => {
            event.preventDefault();
            details.removeAttribute('open');
            const methodologyButton = Array.from(document.querySelectorAll('button')).find(
              (button) => button.textContent.trim().toLowerCase() === 'methodology'
            );
            if (methodologyButton) methodologyButton.click();
            else window.location.href = '?view=methodology#top';
          });
        }
      })();
    </script>
    """
    st.html(header_html, unsafe_allow_javascript=True)
    st.markdown(
        '<div class="main-subtitle">Assessing energy transition opportunities across coal-dependent markets</div>',
        unsafe_allow_html=True,
    )

def render_survey_header() -> None:
    """Render the app identity without links that could influence survey respondents."""
    st.markdown('<span id="top" class="anchor"></span>', unsafe_allow_html=True)
    st.markdown(
        '<div class="main-title-row"><div class="main-title">Coal-to-Clean Jurisdictional Readiness Index 2026</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="main-subtitle">Assessing energy transition opportunities across coal-dependent markets</div>',
        unsafe_allow_html=True,
    )


def render_mobile_desktop_notice() -> None:
    """Advise phone users to reopen the analytical app on a larger screen."""
    st.markdown(
        """
        <div class="mobile-desktop-notice" role="note">
          <strong>Please view the index on a laptop or desktop.</strong>
          The research survey is mobile-friendly, but the index charts, tables and navigation are not optimised for phones.
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_top_navigation(view: str, market_period: str | None = None) -> None:
    with st.container(key="top_nav"):
        nav_columns = st.columns([1.2, 1.68, 2.05, 4.07], gap="small")
        nav_items = [
            ("overall", "Main Index"),
            ("market", "Market Confidence"),
            ("priorities", "Add Your Priorities"),
        ]
        for column, (target, label) in zip(nav_columns[:3], nav_items):
            state = "active" if view == target else "inactive"
            with column:
                with st.container(key=f"nav_{target}_{state}"):
                    if st.button(label, key=f"top_nav_{target}", type="tertiary"):
                        if target == "priorities" and view != "priorities":
                            reset_priority_survey()
                        navigate_to(target, scroll_target="top")

    if view == "market":
        period = market_period or "q1"
        with st.container(key="period_slot"):
            # The empty first column matches the Main Index navigation column,
            # so Q1 begins directly beneath Market Confidence.
            period_columns = st.columns(
                [1.2, 0.9, 0.9, 1.5, 4.5],
                gap="medium",
            )
            with period_columns[1]:
                with st.container(key=f"period_q1_{'active' if period == 'q1' else 'inactive'}"):
                    if st.button("Q1 2026", key="period_q1", type="tertiary"):
                        navigate_to("market", period="q1", scroll_target="top")
            with period_columns[2]:
                with st.container(key=f"period_q2_{'active' if period == 'q2' else 'inactive'}"):
                    if st.button("Q2 2026", key="period_q2", type="tertiary"):
                        navigate_to("market", period="q2", scroll_target="top")
            with period_columns[3]:
                with st.container(key="period_clear"):
                    if st.button("Clear adjustment", key="period_clear_button", type="tertiary"):
                        navigate_to("overall", scroll_target="top")
    else:
        st.markdown('<div class="period-placeholder">Placeholder</div>', unsafe_allow_html=True)


def base_results(index_data: pd.DataFrame) -> pd.DataFrame:
    result = index_data[
        ["Country", "ISO3", "Base index rank", "Base index score", "Operating coal capacity (MW)", *PILLAR_SOURCES]
    ].copy()
    result = result.rename(columns={"Base index rank": "Base rank", "Base index score": "Base score"})
    result["Rank"] = result["Base rank"].astype(int)
    result["Overall score"] = result["Base score"].astype(float)
    result["Rank change"] = 0
    result["Score change"] = 0.0
    return result.sort_values("Rank").reset_index(drop=True)


def adjusted_weights(industry: str, responses: dict[str, int]) -> list[float]:
    slider_multipliers = {1: 0.60, 2: 0.80, 3: 1.00, 4: 1.20, 5: 1.40}
    pillar_responses: list[list[int]] = [[] for _ in range(6)]
    for item in SURVEY_ITEMS:
        for pillar_index in item["pillars"]:
            pillar_responses[pillar_index].append(responses[item["key"]])
    response_values = [sum(values) / len(values) if values else 3 for values in pillar_responses]
    industry_values = INDUSTRY_MULTIPLIERS[industry]
    pre_normalised = [
        base * industry_multiplier * slider_multipliers[int(round(response_value))]
        for base, industry_multiplier, response_value in zip(BASE_WEIGHTS, industry_values, response_values)
    ]
    total = sum(pre_normalised)
    return [value / total for value in pre_normalised]


def calculate_custom_results(index_data: pd.DataFrame, weights: list[float]) -> pd.DataFrame:
    result = index_data[
        ["Country", "ISO3", "Base index rank", "Base index score", "Operating coal capacity (MW)", *PILLAR_SOURCES]
    ].copy()
    result = result.rename(columns={"Base index rank": "Base rank", "Base index score": "Base score"})
    result["Overall score"] = sum(result[column] * weight for column, weight in zip(PILLAR_SOURCES, weights))
    result["Rank"] = result["Overall score"].rank(method="min", ascending=False).astype(int)
    result["Rank change"] = result["Base rank"].astype(int) - result["Rank"]
    result["Score change"] = result["Overall score"] - result["Base score"]
    return result.sort_values(["Rank", "Country"]).reset_index(drop=True)


def ranking_figure(results: pd.DataFrame, show_all: bool, mode: str) -> go.Figure:
    displayed = results.sort_values("Rank").copy()
    if not show_all:
        displayed = displayed.head(10)
    displayed = displayed.sort_values("Rank", ascending=False)

    colour = "#2b5688"
    if mode == "market":
        colour = "#79a6d2"
    elif mode == "custom":
        colour = "#8bb0da"

    overlay_mode = mode in {"market", "custom"}
    customdata = []
    for _, row in displayed.iterrows():
        rank_change = int(row.get("Rank change", 0))
        change_text = "No change" if rank_change == 0 else f"{rank_change:+d} position{'s' if abs(rank_change) != 1 else ''}"
        customdata.append([int(row["Rank"]), change_text, int(row.get("Base rank", row["Rank"]))])

    score_text = displayed["Overall score"].map(lambda value: f"{float(value):.2f}")
    hover_template = (
        "<b>%{y}</b><br>Adjusted rank: %{customdata[0]}<br>"
        "Position change: %{customdata[1]}<extra></extra>"
        if overlay_mode
        else "<b>%{y}</b><br>Rank: %{customdata[0]}<extra></extra>"
    )

    figure = go.Figure()
    figure.add_trace(
        go.Bar(
            x=displayed["Overall score"],
            y=displayed["Country"],
            orientation="h",
            marker_color=colour,
            text=score_text,
            textposition="inside" if overlay_mode else "outside",
            insidetextanchor="start" if overlay_mode else "end",
            textfont=dict(
                family="Montserrat",
                size=12,
                color="#fff5cc" if overlay_mode else "rgba(0,0,0,.86)",
            ),
            cliponaxis=False,
            customdata=customdata,
            hovertemplate=hover_template,
            name="Adjusted score" if overlay_mode else "Overall score",
        )
    )

    if overlay_mode:
        connector_x: list[float | None] = []
        connector_y: list[str | None] = []
        for _, row in displayed.iterrows():
            connector_x.extend([float(row["Base score"]), float(row["Overall score"]), None])
            connector_y.extend([str(row["Country"]), str(row["Country"]), None])
        figure.add_trace(
            go.Scatter(
                x=connector_x,
                y=connector_y,
                mode="lines",
                line=dict(color="rgba(43,86,136,.48)", width=2),
                hoverinfo="skip",
                showlegend=False,
            )
        )
        figure.add_trace(
            go.Scatter(
                x=displayed["Base score"],
                y=displayed["Country"],
                mode="markers",
                customdata=customdata,
                marker=dict(symbol="diamond", size=13, color="#2b5688", line=dict(width=1, color="#ffffff")),
                hovertemplate="<b>%{y}</b><br>Base rank: %{customdata[2]}<br>Base index score: %{x:.2f}<extra></extra>",
                name="Base index",
            )
        )

    maximum = max(100.0, float(displayed["Overall score"].max()) + 8)
    row_height = 42
    chart_height = (len(displayed) * row_height) + 66
    figure.update_layout(
        autosize=False,
        height=chart_height,
        margin=dict(l=280, r=78, t=0, b=66, autoexpand=False),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Montserrat", color="rgba(0,0,0,.90)"),
        bargap=0.22,
        showlegend=overlay_mode,
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1.0,
            xanchor="right",
            x=0.99,
            font=dict(size=11),
            traceorder="normal",
            bgcolor="rgba(0,0,0,0)",
        ),
        hoverlabel=dict(
            bgcolor="#76a7d5",
            bordercolor="#ffffff",
            font=dict(family="Montserrat", size=12, color="#ffffff"),
        ),
        dragmode=False,
        uniformtext=dict(minsize=10, mode="show"),
    )
    figure.update_xaxes(
        title=dict(text="OVERALL SCORE", font=dict(size=14, family="Montserrat"), standoff=18),
        range=[0, maximum],
        fixedrange=True,
        gridcolor="rgba(43,86,136,.11)",
        zeroline=False,
        tickfont=dict(size=11),
    )
    figure.update_yaxes(
        fixedrange=True,
        domain=[0.0, 1.0],
        range=[-0.5, len(displayed) - 0.5],
        tickfont=dict(size=13, family="Montserrat", color="rgba(0,0,0,.88)"),
        automargin=False,
        ticklabelposition="outside",
        ticklabelstandoff=28,
        categoryorder="array",
        categoryarray=displayed["Country"].tolist(),
    )
    return figure

def pillar_figure(data_row: pd.Series, weights: list[float] | None = None) -> go.Figure:
    values = [float(data_row[column]) for column in PILLAR_SOURCES]
    labels = [f"Pillar {index}" for index in range(1, 7)]
    text = [f"{value:.1f}" for value in values]
    figure = go.Figure(
        go.Bar(
            x=labels,
            y=values,
            marker_color="#2b5688" if weights is None else "#8bb0da",
            width=0.38,
            text=text,
            textposition="outside",
            textfont=dict(family="Montserrat", size=12),
            hoverinfo="skip",
        )
    )
    figure.update_layout(
        height=430,
        margin=dict(l=72, r=20, t=20, b=78),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Montserrat", color="rgba(0,0,0,.90)"),
        dragmode=False,
        showlegend=False,
    )
    figure.update_yaxes(
        title=dict(text="PILLAR SCORE", font=dict(size=14, family="Montserrat"), standoff=16),
        range=[0, 105],
        fixedrange=True,
        gridcolor="rgba(43,86,136,.11)",
        zeroline=False,
        tickfont=dict(size=11),
    )
    figure.update_xaxes(
        fixedrange=True,
        tickfont=dict(size=12, family="Montserrat"),
        ticklabelstandoff=16,
    )
    return figure

def descriptor_for_country(
    country: str,
    current_row: pd.Series,
    results: pd.DataFrame,
    index_data: pd.DataFrame,
    custom: bool,
) -> str:
    ordered = results.sort_values("Rank").reset_index(drop=True)
    rank = int(current_row["Rank"])
    score = float(current_row["Overall score"])
    total = len(ordered)
    median_score = float(ordered["Overall score"].median())
    median_gap = score - median_score

    sentences: list[str] = []
    if rank == 1 and total > 1:
        second = ordered.iloc[1]
        lead = score - float(second["Overall score"])
        sentences.append(
            f"{country} leads the index by {lead:.1f} points over {second['Country']} and scores {abs(median_gap):.1f} points above the jurisdiction median."
        )
    elif rank == total and total > 1:
        above = ordered.iloc[-2]
        gap = float(above["Overall score"]) - score
        sentences.append(
            f"{country} trails the next-ranked jurisdiction by {gap:.1f} points and sits {abs(median_gap):.1f} points below the median."
        )
    else:
        above = ordered.iloc[rank - 2] if rank > 1 else None
        below = ordered.iloc[rank] if rank < total else None
        neighbours = []
        if above is not None:
            neighbours.append((float(above["Overall score"]) - score, str(above["Country"])))
        if below is not None:
            neighbours.append((score - float(below["Overall score"]), str(below["Country"])))
        nearest_gap, nearest_country = min(neighbours, key=lambda item: item[0])
        direction = "above" if median_gap >= 0 else "below"
        sentences.append(
            f"{country} sits {abs(median_gap):.1f} points {direction} the median and is only {nearest_gap:.1f} points from {nearest_country}."
        )

    data_row = index_data.loc[index_data["Country"] == country].iloc[0]
    pillar_scores = pd.Series({column: float(data_row[column]) for column in PILLAR_SOURCES})
    pillar_medians = index_data[PILLAR_SOURCES].median()
    above_count = int((pillar_scores > pillar_medians).sum())
    score_range = float(pillar_scores.max() - pillar_scores.min())

    if score_range <= 20:
        sentences.append(
            f"Its readiness profile is comparatively balanced, with scores above the peer median in {above_count} of six pillars."
        )
    else:
        sentences.append(
            f"Its readiness profile is uneven, with scores above the peer median in {above_count} of six pillars and a {score_range:.1f}-point spread across pillars."
        )

    sorted_scores = pillar_scores.sort_values(ascending=False)
    strongest_gap = float(sorted_scores.iloc[0] - sorted_scores.iloc[1])
    weakest_gap = float(sorted_scores.iloc[-2] - sorted_scores.iloc[-1])
    if strongest_gap >= 10:
        source = sorted_scores.index[0]
        sentences.append(
            f"Its clearest differentiator is **{PILLAR_FULL[source]}**, which is {strongest_gap:.1f} points ahead of its next-best pillar."
        )
    elif weakest_gap >= 10:
        source = sorted_scores.index[-1]
        sentences.append(
            f"Its clearest constraint is **{PILLAR_FULL[source]}**, which is {weakest_gap:.1f} points behind its next-weakest pillar."
        )

    if custom:
        rank_change = int(current_row["Rank change"])
        score_change = float(current_row["Score change"])
        if rank_change > 0:
            sentences.append(
                f"Under the selected priorities, it rises {rank_change} position{'s' if rank_change != 1 else ''} and its score changes by {score_change:+.1f} points."
            )
        elif rank_change < 0:
            sentences.append(
                f"Under the selected priorities, it falls {abs(rank_change)} position{'s' if rank_change != -1 else ''} and its score changes by {score_change:+.1f} points."
            )
        else:
            sentences.append(
                f"Under the selected priorities, its rank is unchanged and its score changes by {score_change:+.1f} points."
            )

    return " ".join(sentences)


def bold_markdown_to_html(text: str) -> str:
    return re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)


def priority_considerations(country: str, index_data: pd.DataFrame) -> list[tuple[str, str]]:
    row = index_data.loc[index_data["Country"] == country].iloc[0]
    scores = pd.Series({column: float(row[column]) for column in PILLAR_SOURCES}).sort_values()
    selected = list(scores.index[:2])
    return [(PILLAR_FULL[column], RECOMMENDATIONS[column]) for column in selected]


def current_result_query(view: str, market_period: str | None, custom_version: dict[str, Any] | None) -> str:
    parts = [f"view={view}"]
    if market_period:
        parts.append(f"period={market_period}")
    if custom_version:
        parts.append(f"result=preview_{custom_version['id']}")
    return "?" + "&".join(parts)


def render_html_table(
    table: pd.DataFrame,
    current_query: str,
    height: int = 560,
) -> None:
    rows = []
    for _, row in table.iterrows():
        cells = []
        for column in table.columns:
            value = row[column]
            if pd.isna(value):
                rendered = ""
            elif isinstance(value, (float, int)) and column not in {"Rank", "Change"}:
                rendered = f"{float(value):.2f}"
            elif column in {"Rank", "Change"}:
                rendered = str(int(value)) if float(value).is_integer() else str(value)
            else:
                rendered = html.escape(str(value))
            class_name = "rank-cell" if column == "Rank" else ""
            cells.append(f'<td class="{class_name}" data-col="{html.escape(column)}">{rendered}</td>')
        country = quote(str(row["Country"]))
        target = f"{current_query}&country={country}#country-assessment"
        rows.append(f'<tr onclick="window.location.href=\'{target}\'">{"".join(cells)}</tr>')

    headers = "".join(
        f'<th data-col="{html.escape(column)}">{html.escape(column)}</th>' for column in table.columns
    )
    column_options = "".join(
        f'<label><input type="checkbox" checked data-toggle="{html.escape(column)}"><span>{html.escape(column)}</span></label>'
        for column in table.columns
    )

    col_widths = []
    for column in table.columns:
        if column == "Rank":
            width = 62
        elif column == "Country":
            width = 200
        elif column == "Overall Score":
            width = 100
        elif column == "Change":
            width = 105
        else:
            width = 145
        col_widths.append(f'<col data-col="{html.escape(column)}" style="width:{width}px">')

    component_html = f"""
    <div class="readiness-table-component" id="table-component">
      <style>
        .readiness-table-component {{ height:{height}px; display:flex; flex-direction:column; font-family:Montserrat,sans-serif; color:rgba(0,0,0,.90); }}
        .readiness-table-component * {{ box-sizing:border-box; }}
        .readiness-table-toolbar {{ height:64px; flex:0 0 64px; display:flex; align-items:flex-end; justify-content:flex-end; gap:8px; padding-top:24px; margin-bottom:10px; background:transparent; position:relative; overflow:visible; z-index:100; }}
        .readiness-tool {{ position:relative; }}
        .readiness-tool summary {{ width:38px; height:34px; border:1px solid rgba(43,86,136,.42); border-radius:7px; background:#ffffff; display:flex; align-items:center; justify-content:center; cursor:pointer; list-style:none; color:#2b5688 !important; opacity:1; }}
        .readiness-tool summary::-webkit-details-marker {{ display:none; }}
        .readiness-tool summary:hover {{ background:#fff5cc; }}
        .columns-glyph {{ display:block; color:#2b5688; font-size:22px; line-height:1; font-family:Arial,sans-serif; font-weight:700; transform:translateY(-1px); }}
        .readiness-column-menu {{ position:absolute; right:0; top:39px; z-index:2000; width:min(760px, 92vw); background:white; border:1px solid rgba(43,86,136,.16); border-radius:10px; padding:14px 16px; box-shadow:0 12px 30px rgba(43,86,136,.18); display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); column-gap:24px; }}
        .readiness-column-menu label {{ display:flex; align-items:center; gap:10px; font-size:12px; line-height:1.25; margin:8px 0; white-space:nowrap; }}
        .readiness-column-menu input[type="checkbox"] {{ width:16px; height:16px; accent-color:#ffe680; flex:0 0 auto; }}
        .readiness-table-shell {{ flex:1; min-height:0; border:1px solid rgba(43,86,136,.14); border-radius:10px; overflow:hidden; background:white; }}
        .readiness-table-scroll {{ height:100%; overflow:auto; }}
        .readiness-table-component table {{ border-collapse:collapse; width:100%; min-width:1232px; table-layout:fixed; }}
        .readiness-table-component thead th {{ position:sticky; top:0; z-index:3; height:76px; padding:10px 11px; text-align:left; vertical-align:middle; white-space:normal; line-height:1.25; font-size:13px; font-weight:600; background:#fffae6; border-right:1px solid rgba(43,86,136,.10); border-bottom:1px solid rgba(43,86,136,.14); }}
        .readiness-table-component tbody td {{ height:39px; padding:8px 11px; font-size:13px; font-weight:500; border-right:1px solid rgba(43,86,136,.08); border-bottom:1px solid rgba(43,86,136,.08); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
        .readiness-table-component tbody tr {{ cursor:pointer; }}
        .readiness-table-component tbody tr:nth-child(even) {{ background:#fffdf3; }}
        .readiness-table-component tbody tr:hover {{ background:#fff5cc; }}
        .readiness-table-component .rank-cell {{ text-align:center; }}
        .readiness-tooltip {{ position:absolute; left:50%; bottom:42px; transform:translateX(-50%); padding:5px 8px; background:rgba(0,0,0,.78); color:white; border-radius:5px; font-size:10px; white-space:nowrap; opacity:0; visibility:hidden; pointer-events:none; transition:opacity .12s ease, visibility .12s ease; }}
        .readiness-tooltip::after {{ content:""; position:absolute; left:50%; top:100%; transform:translateX(-50%); border:5px solid transparent; border-top-color:rgba(0,0,0,.78); }}
        .readiness-tool:hover .readiness-tooltip, .readiness-tool:focus-within .readiness-tooltip {{ opacity:1; visibility:visible; }}
      </style>
      <div class="readiness-table-toolbar">
        <details class="readiness-tool" id="column-picker">
          <summary aria-label="Choose columns" title="Choose columns"><span class="columns-glyph" aria-hidden="true">▥</span></summary>
          <span class="readiness-tooltip" role="tooltip">Choose columns</span>
          <div class="readiness-column-menu">{column_options}</div>
        </details>
      </div>
      <div class="readiness-table-shell">
        <div class="readiness-table-scroll">
          <table>
            <colgroup>{''.join(col_widths)}</colgroup>
            <thead><tr>{headers}</tr></thead>
            <tbody>{''.join(rows)}</tbody>
          </table>
        </div>
      </div>
    </div>
    <script>
      document.querySelectorAll('#table-component [data-toggle]').forEach(function(box) {{
        box.addEventListener('change', function() {{
          const col = this.getAttribute('data-toggle');
          document.querySelectorAll('#table-component [data-col="' + CSS.escape(col) + '"]').forEach(function(cell) {{
            cell.style.display = box.checked ? '' : 'none';
          }});
        }});
      }});
      document.addEventListener('click', function(event) {{
        const picker = document.getElementById('column-picker');
        if (picker && picker.open && !picker.contains(event.target)) picker.removeAttribute('open');
      }});
    </script>
    """
    st.html(component_html, unsafe_allow_javascript=True)

def render_chart(
    figure: go.Figure,
    filename: str,
    key: str,
) -> None:
    st.plotly_chart(
        figure,
        width="stretch",
        theme=None,
        key=key,
        config={
            "displayModeBar": True,
            "displaylogo": False,
            "scrollZoom": False,
            "responsive": True,
            "showTips": True,
            "modeBarButtonsToRemove": [
                "zoom2d",
                "pan2d",
                "select2d",
                "lasso2d",
                "zoomIn2d",
                "zoomOut2d",
                "autoScale2d",
                "resetScale2d",
            ],
            "toImageButtonOptions": {"format": "png", "filename": filename, "scale": 2},
        },
    )


def render_chart_explainer(view: str, market_period: str | None, mode: str) -> None:
    """Add concise interpretation below the ranking chart."""
    scope = (
        "This index assesses jurisdictions with grid-connected coal-fired power plants currently in operation, <br>based on Global Energy Monitor’s Global Coal Plant Tracker data (January 2026 release)."
    )
    if view == "market":
        quarter = "Q1 2026" if market_period == "q1" else "Q2 2026"
        st.markdown(
            f"""
            <div class="chart-explainer">
              <p>The {quarter} overlay changes the relative emphasis placed on existing readiness indicators according to market evidence observed during the quarter. The same adjusted weights are applied across all jurisdictions.</p>
              <p>{scope}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    extra = ""
    if mode == "custom":
        extra = (
            "<p>The adjusted ranking reflects the priorities and industry profile entered in the survey and remains separate from the published index and market-confidence overlays.</p>"
            '<p class="research-survey-invite"><a href="?r=c2c26-7k4m9#top">Complete the research survey on coal-to-clean projects with carbon credits</a></p>'
        )
    st.markdown(
        f'<div class="chart-explainer"><p>{scope}</p>{extra}</div>',
        unsafe_allow_html=True,
    )


@st.fragment
def render_ranking_chart_fragment(
    results: pd.DataFrame,
    view: str,
    mode: str,
    custom_id: str | None,
    market_period: str | None,
) -> None:
    show_all_key = f"show_all_{view}_{custom_id or market_period or 'base'}"
    if show_all_key not in st.session_state:
        st.session_state[show_all_key] = False
    show_all = bool(st.session_state[show_all_key])

    if view == "market":
        label = "Q1 2026" if market_period == "q1" else "Q2 2026"
        chart_title = f"{label}: Market Confidence-Adjusted Country Ranking"
        title_class = "section-title"
    elif mode == "custom":
        chart_title = "Your Priority-Adjusted Country Ranking"
        title_class = "section-title"
    else:
        chart_title = "2026 Country Ranking" if show_all else "Top 10 Country Ranking"
        title_class = "section-title base-chart-title"

    with st.container(key="chart_header_slot"):
        st.markdown(f'<div class="{title_class}">{chart_title}</div>', unsafe_allow_html=True)
        show_all = st.toggle("Show all countries", key=show_all_key)

    with st.container(key="ranking_chart_slot"):
        render_chart(
            ranking_figure(results, show_all=show_all, mode=mode),
            "readiness_index_country_ranking",
            f"ranking_chart_{view}_{custom_id or market_period or 'base'}",
        )
        render_chart_explainer(view, market_period, mode)


def get_visible_previews() -> list[dict[str, Any]]:
    saved_ids = set(st.session_state.saved_preview_ids)
    visible = [preview for preview in st.session_state.previews if preview["id"] in saved_ids]
    current_id = st.session_state.current_preview_id
    if current_id and current_id not in saved_ids:
        current = next((preview for preview in st.session_state.previews if preview["id"] == current_id), None)
        if current:
            visible.append(current)
    return sorted(visible, key=lambda item: item["number"])


def resolve_custom_version() -> dict[str, Any] | None:
    result_param = query_value("result", "")
    if not result_param.startswith("preview_"):
        return None
    preview_id = result_param.replace("preview_", "", 1)
    match = next((preview for preview in st.session_state.previews if preview["id"] == preview_id), None)
    if match:
        st.session_state.current_preview_id = match["id"]
    return match

def render_result_selector(custom_version: dict[str, Any] | None) -> dict[str, Any] | None:
    visible = get_visible_previews()
    if not visible:
        return custom_version
    labels = ["Main index"] + [f"Customised view {preview['number']}" for preview in visible]
    current_label = "Main index"
    if custom_version:
        current_label = f"Customised view {custom_version['number']}"
    if "result_view_selector" not in st.session_state or st.session_state.result_view_selector not in labels:
        st.session_state.result_view_selector = current_label
    selected = st.segmented_control(
        "Result view",
        labels,
        key="result_view_selector",
        label_visibility="collapsed",
    )
    if not selected or selected == current_label:
        return custom_version
    if selected == "Main index":
        st.session_state.current_preview_id = None
        if "result" in st.query_params:
            del st.query_params["result"]
        st.rerun()
    number = int(selected.split()[-1])
    chosen = next(preview for preview in visible if preview["number"] == number)
    st.session_state.current_preview_id = chosen["id"]
    st.query_params["result"] = f"preview_{chosen['id']}"
    st.rerun()
    return custom_version


def display_results(
    index_data: pd.DataFrame,
    results: pd.DataFrame,
    view: str,
    custom_version: dict[str, Any] | None = None,
    market_period: str | None = None,
) -> None:
    custom = custom_version is not None
    mode = "custom" if custom else ("market" if view == "market" else "base")

    with st.container(key="ranking_section"):
        render_top_navigation(view, market_period)
        if custom_version:
            results = custom_version["results"]

        render_ranking_chart_fragment(
            results=results,
            view=view,
            mode=mode,
            custom_id=custom_version["id"] if custom_version else None,
            market_period=market_period,
        )

        if custom:
            render_custom_follow_up(index_data, custom_version)

    st.markdown('<div class="section-gap"></div>', unsafe_allow_html=True)
    with st.container(key="table_section"):
        st.markdown('<span id="pillar-scores" class="anchor"></span>', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Jurisdiction Rankings and Pillar Scores</div>', unsafe_allow_html=True)

        table = results.sort_values("Rank").copy()
        table_columns = ["Rank", "Country", "Overall score"]
        if custom:
            table_columns.append("Rank change")
        table_columns.extend(PILLAR_SOURCES)
        table = table[table_columns].rename(columns=PILLAR_TABLE)
        table = table.rename(columns={"Overall score": "Overall Score", "Rank change": "Change"})
        query = current_result_query(view, market_period, custom_version)
        render_html_table(table, query)

    st.markdown('<div class="section-gap"></div>', unsafe_allow_html=True)
    with st.container(key="assessment_section"):
        st.markdown('<span id="country-assessment" class="anchor"></span>', unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="margin-bottom:20px;">Country Assessment</div>', unsafe_allow_html=True)

        countries = sorted(results["Country"].tolist())
        query_country = query_value("country", "Chile")
        preferred = query_country if query_country in countries else st.session_state.selected_country
        if preferred not in countries:
            preferred = "Chile" if "Chile" in countries else countries[0]
        st.session_state.selected_country = preferred
        if "country_selector" not in st.session_state or st.session_state.country_selector not in countries:
            st.session_state.country_selector = preferred

        left, right = st.columns([1.05, 1.85], gap="large")
        with left:
            with st.container(key="country_picker"):
                selected_country = st.selectbox(
                    "Search jurisdiction",
                    countries,
                    key="country_selector",
                    width="stretch",
                )
            st.session_state.selected_country = selected_country
            current_row = results.loc[results["Country"] == selected_country].iloc[0]
            st.markdown(
                f"""
                <div class="metric-grid">
                  <div class="metric-card"><div class="metric-label">Overall rank</div><div class="metric-value">{int(current_row['Rank'])} of {len(results)}</div></div>
                  <div class="metric-card"><div class="metric-label">Overall score</div><div class="metric-value">{float(current_row['Overall score']):.2f}</div></div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            assessment = descriptor_for_country(selected_country, current_row, results, index_data, custom)
            st.markdown('<div class="assessment-heading">Assessment</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="assessment-copy">{bold_markdown_to_html(assessment)}</div>', unsafe_allow_html=True)

        with right:
            data_row = index_data.loc[index_data["Country"] == selected_country].iloc[0]
            weights = custom_version["weights"] if custom else None
            render_chart(
                pillar_figure(data_row, weights),
                f"{selected_country}_pillar_scores",
                f"pillar_chart_{selected_country}_{custom_version['id'] if custom_version else 'base'}",
            )

        with st.container(key="assessment_bottom_row"):
            bottom_left, bottom_right = st.columns([1.05, 1.85], gap="large")
            with bottom_left:
                with st.container(key="priority_panel"):
                    st.markdown('<div class="priority-heading">Priority considerations</div>', unsafe_allow_html=True)
                    priority_paragraphs = "".join(
                        f'<p><strong>{html.escape(label)}</strong> – {html.escape(recommendation)}</p>'
                        for label, recommendation in priority_considerations(selected_country, index_data)
                    )
                    st.markdown(
                        f'<div class="priority-card">{priority_paragraphs}</div>',
                        unsafe_allow_html=True,
                    )

            with bottom_right:
                with st.container(key="pillar_explanation_panel"):
                    legend_items = "".join(
                        (
                            '<div class="pillar-definition">'
                            f'<strong>{html.escape(PILLAR_FULL[source])}</strong>'
                            f'<p>{html.escape(description)}</p>'
                            '</div>'
                        )
                        for source, description in zip(PILLAR_SOURCES, PILLAR_DESCRIPTIONS)
                    )
                    st.markdown(f'<div class="pillar-legend-box">{legend_items}</div>', unsafe_allow_html=True)

        st.markdown(
            '<div class="assessment-footnote">This commentary is generated from relative index and pillar scores. It is not a substitute for transaction-specific due diligence.</div>',
            unsafe_allow_html=True,
        )


def report_versions_for_export(current_version: dict[str, Any]) -> list[dict[str, Any]]:
    """The current customised view is available for export as soon as it is created."""
    return [current_version]


def begin_custom_action(action: str, current_version: dict[str, Any]) -> None:
    st.session_state.action_notice = None
    st.session_state.action_in_progress = action
    st.session_state.action_version_id = current_version["id"]


def render_action_progress(label: str) -> None:
    st.markdown(
        (
            '<div class="action-progress-wrap">'
            '<div class="action-progress-spinner" aria-hidden="true"></div>'
            f'<div>{html.escape(label)}</div>'
            '</div>'
        ),
        unsafe_allow_html=True,
    )


def prepare_pdf(index_data: pd.DataFrame, current_version: dict[str, Any]) -> tuple[bool, str]:
    try:
        st.session_state.cached_pdf = build_versions_pdf(
            base_results(index_data),
            report_versions_for_export(current_version),
        )
        st.session_state.cached_pdf_version_id = current_version["id"]
        return True, "PDF ready."
    except Exception as exc:
        st.session_state.cached_pdf = None
        st.session_state.cached_pdf_version_id = None
        return False, "The PDF could not be prepared. " + str(exc)


def perform_custom_action(index_data: pd.DataFrame, current_version: dict[str, Any]) -> None:
    """Complete a queued PDF preparation or email action."""
    action = st.session_state.action_in_progress
    version_id = st.session_state.action_version_id
    if not action or version_id != current_version["id"]:
        return

    time.sleep(0.18)

    if action == "prepare_download":
        ok, message = prepare_pdf(index_data, current_version)
        st.session_state.action_notice = (
            "Your PDF is ready. Click Download PDF once more to save it."
            if ok
            else message
        )
        st.session_state.action_notice_type = "success" if ok else "error"
    elif action == "email":
        ok, message = send_results_email(index_data, current_version)
        st.session_state.action_notice = message
        st.session_state.action_notice_type = "success" if ok else "error"

    st.session_state.action_in_progress = None
    st.session_state.action_version_id = None
    st.rerun()


def render_custom_follow_up(index_data: pd.DataFrame, current_version: dict[str, Any]) -> None:
    with st.container(key="custom_actions"):
        if st.session_state.logging_notice:
            st.markdown(
                f'<div class="action-message error">{html.escape(st.session_state.logging_notice)}</div>',
                unsafe_allow_html=True,
            )
        if st.session_state.action_notice:
            notice_type = st.session_state.action_notice_type or "warning"
            st.markdown(
                f'<div class="action-message {html.escape(notice_type)}">{html.escape(st.session_state.action_notice)}</div>',
                unsafe_allow_html=True,
            )

        left_space, download_col, gap, email_col, right_space = st.columns(
            [2.8, 1.7, 0.5, 1.7, 2.8], gap="small"
        )
        del left_space, gap, right_space

        cached_pdf_ready = (
            st.session_state.cached_pdf is not None
            and st.session_state.cached_pdf_version_id == current_version["id"]
        )
        download_authorised = st.session_state.download_authorised_version_id == current_version["id"]

        with download_col:
            if download_authorised and cached_pdf_ready:
                st.download_button(
                    "Download PDF",
                    data=st.session_state.cached_pdf,
                    file_name="Coal-to-Clean Jurisdictional Readiness Index 2026.pdf",
                    mime="application/pdf",
                    on_click="ignore",
                    key=f"download_{current_version['id']}",
                    width="stretch",
                )
            elif st.button("Download PDF", key=f"gate_download_{current_version['id']}", width="stretch"):
                st.session_state.action_notice = None
                st.session_state.gate_action = "download"
                st.session_state.gate_version_id = current_version["id"]
                st.session_state.gate_processing = False
                st.rerun()

        with email_col:
            if st.button("Email results", key=f"email_{current_version['id']}", width="stretch"):
                st.session_state.action_notice = None
                st.session_state.gate_action = "email"
                st.session_state.gate_version_id = current_version["id"]
                st.session_state.gate_processing = False
                st.rerun()



def valid_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value.strip()))


def send_results_email(index_data: pd.DataFrame, current_version: dict[str, Any]) -> tuple[bool, str]:
    if not backend_enabled():
        return False, "Email delivery is not connected yet. Please download the PDF instead."

    contact = st.session_state.contact
    selected_versions = report_versions_for_export(current_version)
    pdf_content = build_versions_pdf(base_results(index_data), selected_versions)
    attachments = [
        encode_attachment("Coal-to-Clean Jurisdictional Readiness Index 2026.pdf", "application/pdf", pdf_content)
    ]

    ok, message = post_backend(
        "send_email",
        {
            "session_id": st.session_state.session_id,
            "requested_at": singapore_now(),
            "name": contact["name"],
            "organisation": contact["organisation"],
            "email": contact["email"],
            "versions_requested": [f"Customised view {version['number']}" for version in selected_versions],
            "attachments": attachments,
        },
    )
    if ok:
        return True, "Your results have been emailed!"
    if "Unauthorised" in message or "unauthorised" in message:
        return False, "The Google Sheets connection needs to be reauthorised before email delivery can work."
    return False, "Email delivery was not completed. Please download the PDF instead"

def lookup_preview(preview_id: str | None) -> dict[str, Any] | None:
    if not preview_id:
        return None
    return next((preview for preview in st.session_state.previews if preview["id"] == preview_id), None)


def log_profile_details(reason: str) -> tuple[bool, str]:
    contact = st.session_state.contact
    return post_backend(
        "log_profile",
        {
            "session_id": st.session_state.session_id,
            "captured_at": singapore_now(),
            "name": contact.get("name", ""),
            "organisation": contact.get("organisation", ""),
            "email": contact.get("email", ""),
            "reason": reason,
            "industry": st.session_state.profile_industry,
            "role": st.session_state.profile_role,
            "started_at": st.session_state.session_started_at,
        },
        timeout=12,
    )


@st.dialog("A small request", width="small")
def download_gate_dialog(index_data: pd.DataFrame) -> None:
    current_version = lookup_preview(st.session_state.gate_version_id) or resolve_custom_version()
    if current_version is None:
        st.error("No customised view is available for this action.")
        return

    if st.session_state.gate_processing:
        st.markdown(
            '<div class="gate-inline-progress"><div class="gate-progress-spinner" aria-hidden="true"></div><div>Generating your customised PDF</div></div>',
            unsafe_allow_html=True,
        )
        time.sleep(0.18)
        if st.session_state.profile_capture_pending:
            ok, message = log_profile_details("download")
            st.session_state.session_logged = st.session_state.session_logged or ok
            st.session_state.logging_notice = None if ok else "Your PDF can continue, but Google Sheets logging did not complete. " + message
            st.session_state.profile_capture_pending = False

        ok, message = prepare_pdf(index_data, current_version)
        st.session_state.gate_processing = False
        st.session_state.gate_action = None
        st.session_state.gate_version_id = None
        if ok:
            st.session_state.download_authorised_version_id = current_version["id"]
            st.session_state.action_notice = "Your PDF is ready. Click Download PDF to save it."
            st.session_state.action_notice_type = "success"
        else:
            st.session_state.action_notice = message
            st.session_state.action_notice_type = "error"
        st.rerun()

    st.markdown(
        '<div class="gate-copy">Please share your details as it helps us understand which market participants use the index.</div>',
        unsafe_allow_html=True,
    )
    contact = st.session_state.contact
    name = st.text_input("Name (optional)", value=contact.get("name", ""), key="download_name")
    organisation = st.text_input("Organisation (optional)", value=contact.get("organisation", ""), key="download_organisation")
    email = st.text_input("Email (optional)", value=contact.get("email", ""), key="download_email")
    st.markdown(
        '<div class="gate-fine-print">Details are used only for the stated research purpose and to fulfil requested services. They are not used for marketing.</div>',
        unsafe_allow_html=True,
    )

    left, centre, right = st.columns([1.5, 2.0, 1.5])
    del left, right
    with centre:
        with st.container(key="gate_primary_button"):
            share = st.button("OK", key="download_share", width="stretch")
    with st.container(key="gate_decline_button"):
        decline = st.button("No thanks, I just want the PDF", key="download_decline", type="tertiary", width="stretch")

    if share:
        if not any([name.strip(), organisation.strip(), email.strip()]):
            st.error("Enter at least one detail, or select No thanks, I just want the PDF.")
            return
        if email.strip() and not valid_email(email):
            st.error("Enter a valid email address or leave the email field blank.")
            return
        st.session_state.contact = {"name":name.strip(), "organisation":organisation.strip(), "email":email.strip()}
        st.session_state.profile_complete = bool(email.strip() and valid_email(email))
        st.session_state.profile_capture_pending = True
        st.session_state.gate_processing = True
        st.rerun()

    if decline:
        st.session_state.profile_capture_pending = False
        st.session_state.gate_processing = True
        st.rerun()


@st.dialog("Please enter your details", width="small")
def email_gate_dialog(index_data: pd.DataFrame) -> None:
    current_version = lookup_preview(st.session_state.gate_version_id) or resolve_custom_version()
    if current_version is None:
        st.error("No customised view is available for this action.")
        return

    if st.session_state.gate_processing:
        st.markdown(
            '<div class="gate-inline-progress"><div class="gate-progress-spinner" aria-hidden="true"></div><div>Sending…</div></div>',
            unsafe_allow_html=True,
        )
        time.sleep(0.18)
        profile_ok, profile_message = log_profile_details("email")
        st.session_state.session_logged = st.session_state.session_logged or profile_ok
        st.session_state.logging_notice = None if profile_ok else "Your email can continue, but Google Sheets logging did not complete. " + profile_message
        ok, message = send_results_email(index_data, current_version)
        st.session_state.gate_processing = False
        st.session_state.gate_action = None
        st.session_state.gate_version_id = None
        st.session_state.action_notice = message
        st.session_state.action_notice_type = "success" if ok else "error"
        st.rerun()

    contact = st.session_state.contact
    name = st.text_input("Name", value=contact.get("name", ""), key="email_name")
    organisation = st.text_input("Organisation (optional)", value=contact.get("organisation", ""), key="email_organisation")
    email = st.text_input("Email", value=contact.get("email", ""), key="email_address")
    st.markdown(
        '<div class="gate-fine-print">Your details are used only to fulfil this request and support the stated research purpose. They are not used for marketing.</div>',
        unsafe_allow_html=True,
    )
    left, centre, right = st.columns([1.5, 2.0, 1.5])
    del left, right
    with centre:
        with st.container(key="gate_primary_button"):
            submitted = st.button("Submit", key="email_submit", width="stretch")

    if submitted:
        if not name.strip():
            st.error("Enter your name.")
            return
        if not valid_email(email):
            st.error("Enter a valid email address.")
            return
        st.session_state.contact = {"name":name.strip(), "organisation":organisation.strip(), "email":email.strip()}
        st.session_state.profile_complete = True
        st.session_state.gate_processing = True
        st.rerun()



def render_overall(index_data: pd.DataFrame) -> None:
    custom_version = resolve_custom_version()
    results = custom_version["results"] if custom_version else base_results(index_data)
    display_results(index_data, results, view="overall", custom_version=custom_version)


def render_market_driver_table(drivers: pd.DataFrame) -> None:
    """Render the public-facing quarterly evidence-driver table."""
    if drivers.empty:
        st.info("No market-confidence drivers are available for this period.")
        return

    ordered = (
        drivers.sort_values("Display order")
        if "Display order" in drivers.columns
        else drivers.copy()
    )

    rows: list[str] = []
    for _, row in ordered.iterrows():
        confidence_effect = row.get(
            "Effect on market confidence",
            row.get("Overlay treatment", row.get("Confidence direction", "")),
        )
        readiness_factor = row.get(
            "Readiness factor affected",
            row.get("Mapped indicator", row.get("Index area affected", "")),
        )

        rows.append(
            "<tr>"
            f"<td><strong>{html.escape(str(row.get('Evidence signal', '')))}</strong></td>"
            f"<td class='market-confidence-effect'>{html.escape(str(confidence_effect))}</td>"
            f"<td>{html.escape(str(row.get('Why this matters', row.get('Why it mattered', ''))))}</td>"
            f"<td>{html.escape(str(readiness_factor))}</td>"
            "</tr>"
        )

    st.markdown(
        """
        <table class="market-driver-table">
          <colgroup>
            <col style="width:22%;">
            <col style="width:26%;">
            <col style="width:30%;">
            <col style="width:22%;">
          </colgroup>
          <thead>
            <tr>
              <th>Evidence signal</th>
              <th>Effect on market confidence</th>
              <th>Why this matters</th>
              <th>Readiness factor affected</th>
            </tr>
          </thead>
          <tbody>
        """
        + "".join(rows)
        + """
          </tbody>
        </table>
        """,
        unsafe_allow_html=True,
    )


def render_market_developments(developments: pd.DataFrame) -> None:
    """Render one representative source link per clustered evidence event, using workbook URLs."""
    cards: list[str] = []

    def clean_text(value: Any) -> str:
        if value is None or pd.isna(value):
            return ""
        rendered = str(value).strip()
        return "" if rendered.lower() in {"none", "nan", "nat"} else rendered

    ordered = (
        developments.sort_values("Display order")
        if "Display order" in developments.columns
        else developments.copy()
    )

    for _, row in ordered.iterrows():
        source = clean_text(row.get("Source", ""))
        date = clean_text(row.get("Publication date", ""))
        region = clean_text(row.get("Country or region", ""))

        source_url = clean_text(row.get("Source URL", ""))

        if source_url and source:
            source_html = (
                f'<a href="{html.escape(source_url, quote=True)}" '
                'target="_blank" rel="noopener noreferrer">'
                f'{html.escape(source)}</a>'
            )
        else:
            source_html = html.escape(source)

        meta_parts = [part for part in [source_html, html.escape(date), html.escape(region)] if part]
        meta_html = " • ".join(meta_parts)

        direction = clean_text(
            row.get(
                "Effect on market confidence",
                row.get("Confidence direction", ""),
            )
        )

        direction_key = direction.casefold()
        if "reducing" in direction_key:
            signal_icon = "↓"
        elif "enhancing" in direction_key:
            signal_icon = "↑"
        elif "mixed" in direction_key:
            signal_icon = "↔"
        else:
            signal_icon = "•"

        cards.append(
            "<div class='market-development-card'>"
            f"<div class='market-development-title'>{html.escape(clean_text(row.get('Development', '')))}</div>"
            f"<div class='market-development-meta'>{meta_html}</div>"
            f"<div class='market-development-summary'>{html.escape(clean_text(row.get('Evidence summary', '')))}</div>"
            f"<div class='market-development-signal'>"
            f"<span class='market-development-signal-icon' aria-hidden='true'>{signal_icon}</span>"
            f"<span>{html.escape(direction)} signal</span>"
            "</div></div>"
        )

    st.markdown(
        "<div class='market-development-grid'>" + "".join(cards) + "</div>",
        unsafe_allow_html=True,
    )


def render_market(
    index_data: pd.DataFrame,
    overlay_data: pd.DataFrame,
    market_drivers: pd.DataFrame,
    market_context: pd.DataFrame,
    market_developments: pd.DataFrame,
) -> None:
    period = query_value("period", "q1")
    if period not in {"q1", "q2"}:
        period = "q1"
    quarter = "Q1 2026" if period == "q1" else "Q2 2026"
    results = market_results(index_data, overlay_data, period)

    with st.container(key="ranking_section"):
        render_top_navigation("market", period)
        render_ranking_chart_fragment(
            results=results,
            view="market",
            mode="market",
            custom_id=None,
            market_period=period,
        )

    current_drivers = market_drivers[market_drivers["Quarter"].astype(str) == quarter].copy()
    current_context = market_context[market_context["Quarter"].astype(str) == quarter].copy()
    current_developments = market_developments[market_developments["Quarter"].astype(str) == quarter].copy()

    st.markdown('<div class="section-gap"></div>', unsafe_allow_html=True)
    with st.container(key="market_drivers_section"):
        st.markdown(
            f'<div class="section-title">What drove {quarter}’s adjustment</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div class="market-intro">The {quarter} evidence shows which market signals strengthened, weakened or produced mixed confidence, why they mattered, and the readiness factors they concerned. Country-level score and rank effects still depend on each jurisdiction’s underlying performance on those factors.</div>',
            unsafe_allow_html=True,
        )
        render_market_driver_table(current_drivers)
        if not current_context.empty:
            items = "".join(
                f"<li><strong>{html.escape(str(row['Evidence signal']))}</strong> — {html.escape(str(row['Why it was not scored directly']))}</li>"
                for _, row in current_context.sort_values("Display order").iterrows()
            )
            st.markdown(
                '<div class="market-context-copy assessment-copy">'
                '<div class="market-context-heading">Important evidence not directly scored</div>'
                '<p>These signals were retained as contextual findings because the final index does not contain a sufficiently direct national indicator.</p>'
                f'<ul class="market-context-list">{items}</ul>'
                '</div>',
                unsafe_allow_html=True,
            )

    st.markdown('<div class="section-gap"></div>', unsafe_allow_html=True)
    with st.container(key="market_developments_section"):
        st.markdown(
            f'<div class="section-title">Key developments scored in {quarter}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="market-intro">Representative evidence events are shown below. The arrow summarises the overall direction of each development; mixed means the same event contained both confidence-enhancing and confidence-reducing implications. Duplicate and syndicated reporting was clustered so repeated coverage did not count as separate market signals.</div>',
            unsafe_allow_html=True,
        )
        render_market_developments(current_developments)
        with st.container(key="market_main_link"):
            if st.button("Explore individual jurisdictions and pillar results in the Main Index", key=f"market_to_main_{period}", type="tertiary"):
                navigate_to("overall", scroll_target="top")


def process_priority_submission(index_data: pd.DataFrame, submission: dict[str, Any]) -> None:
    industry = submission["industry"]
    role = submission["role"]
    responses = submission["responses"]

    st.session_state.profile_industry = industry
    st.session_state.profile_role = role

    weights = adjusted_weights(industry, responses)
    results = calculate_custom_results(index_data, weights)
    preview_id = uuid.uuid4().hex[:10]
    submitted_at = singapore_now()
    version = {
        "id": preview_id,
        "number": 1,
        "submitted_at": submitted_at,
        "industry": industry,
        "role": role,
        "responses": responses,
        "weights": weights,
        "results": results,
    }
    st.session_state.previews = [version]
    st.session_state.saved_preview_ids = [preview_id]
    st.session_state.current_preview_id = preview_id
    st.session_state.download_authorised_version_id = None
    st.session_state.cached_pdf = None
    st.session_state.cached_pdf_version_id = None

    rankings_payload = [
        {
            "Country": str(row["Country"]),
            "ISO3": str(row["ISO3"]),
            "Base rank": int(row["Base rank"]),
            "Base score": round(float(row["Base score"]), 6),
            "Adjusted rank": int(row["Rank"]),
            "Adjusted score": round(float(row["Overall score"]), 6),
            "Rank change": int(row["Rank change"]),
        }
        for _, row in results.sort_values("Rank").iterrows()
    ]
    logging_ok, logging_message = post_backend(
        "log_submission",
        {
            "session": {
                "session_id": st.session_state.session_id,
                "started_at": st.session_state.session_started_at,
                "industry": industry,
                "role": role,
            },
            "version": {
                "session_id": st.session_state.session_id,
                "version": 1,
                "submitted_at": submitted_at,
                "industry": industry,
                "role": role,
                "survey_responses": responses,
                "survey_labels": {item["key"]: item["label"] for item in SURVEY_ITEMS},
                "weights": [
                    {
                        "pillar_number": i + 1,
                        "pillar": PILLAR_SHORT[i],
                        "base_weight": BASE_WEIGHTS[i],
                        "adjusted_weight": weights[i],
                    }
                    for i in range(6)
                ],
                "rankings": rankings_payload,
            },
        },
        timeout=15,
    )
    st.session_state.session_logged = logging_ok
    st.session_state.logging_notice = None if logging_ok else (
        "Your ranking was created, but usage logging did not complete. " + logging_message
    )

    st.session_state.pending_priority_submission = None
    st.query_params["view"] = "overall"
    st.query_params["result"] = f"preview_{preview_id}"
    st.session_state.scroll_target = "top"
    st.session_state.scroll_to_top = True
    st.rerun()


def render_priorities(index_data: pd.DataFrame) -> None:
    existing = st.session_state.previews[0] if st.session_state.previews else None
    if existing is not None:
        st.query_params["view"] = "overall"
        st.query_params["result"] = f"preview_{existing['id']}"
        st.session_state.scroll_target = "top"
        st.session_state.scroll_to_top = True
        st.rerun()

    industries = sorted((name for name in INDUSTRY_MULTIPLIERS if name != "Other"), key=str.casefold)
    industry_options = ["Select your industry", *industries, "Other"]
    role_options = ["Select your role", *ROLE_OPTIONS]

    if st.session_state.profile_industry not in industry_options:
        st.session_state.profile_industry = "Select your industry"
    if st.session_state.profile_role not in role_options:
        st.session_state.profile_role = "Select your role"
    if "industry_widget" not in st.session_state or st.session_state.industry_widget not in industry_options:
        st.session_state.industry_widget = st.session_state.profile_industry
    if "role_widget" not in st.session_state or st.session_state.role_widget not in role_options:
        st.session_state.role_widget = st.session_state.profile_role

    with st.container(key="priorities_section"):
        render_top_navigation("priorities")
        st.markdown('<div class="section-title">Set Your Market Priorities</div>', unsafe_allow_html=True)
        st.markdown(
            """
            <div class="assessment-copy" style="max-width:980px; font-size:16px;">
            <p>The published index highlights where coal-to-clean transition opportunities appear most credible, executable and investable across jurisdictions. Different organisations, however, weigh policy certainty, market maturity, financial conditions and social safeguards differently.</p>
            <p>Adjusting the weights lets you apply your own decision criteria to the base index. This reveals which jurisdictions rise or fall under your priorities, helping focus further due diligence, market entry and engagement.</p>
            </div>
            <div class="priority-form-space"></div>
            """,
            unsafe_allow_html=True,
        )

        with st.form("priority_form"):
            col1, col2 = st.columns(2, gap="large")
            with col1:
                with st.container(key="industry_picker"):
                    industry = st.selectbox("Your industry", industry_options, key="industry_widget")
            with col2:
                with st.container(key="role_picker"):
                    role = st.selectbox("Your role", role_options, key="role_widget")
            st.markdown('<div class="priority-dropdown-gap"></div>', unsafe_allow_html=True)
            st.markdown("**When assessing a coal-to-clean opportunity, how important are the following?**")
            st.markdown(
                '<div class="priority-scale-note"><span>1 = Less important</span><span>5 = More important</span></div>',
                unsafe_allow_html=True,
            )

            responses: dict[str, int] = {}
            for item in SURVEY_ITEMS:
                st.markdown(f'<div class="priority-label">{html.escape(item["label"])}</div>', unsafe_allow_html=True)
                st.markdown(
                    f'<div class="priority-description">{html.escape(item["description"])}</div>',
                    unsafe_allow_html=True,
                )
                slider_key = f"priority_slider_{item['key']}"
                if slider_key not in st.session_state:
                    st.session_state[slider_key] = 3
                responses[item["key"]] = st.slider(
                    item["label"], min_value=1, max_value=5, step=1, key=slider_key, label_visibility="collapsed"
                )

            applied = st.form_submit_button("Apply priorities")

        if not applied:
            return
        if industry == "Select your industry" or role == "Select your role":
            st.error("Select both your industry and role before applying your priorities")
            return

        st.session_state.profile_industry = industry
        st.session_state.profile_role = role
        submission = {
            "industry": industry,
            "role": role,
            "responses": dict(responses),
        }
        with st.spinner("Updating your ranking...", show_time=False):
            time.sleep(0.7)
            process_priority_submission(index_data, submission)




def validation_more_info_html(*, compact: bool = False) -> str:
    summary = "ⓘ" if compact else "More about coal-to-clean projects and transition credits"
    trigger_class = "survey-info-icon-trigger" if compact else "survey-info-link-trigger"
    popover_id = "survey-more-info-scenario" if compact else "survey-more-info-intro"
    title = "More about coal-to-clean projects and transition credits"
    return f"""
    <button
      type="button"
      class="survey-info-trigger {trigger_class}"
      popovertarget="{popover_id}"
      aria-label="{title}"
      title="{title}"
    >{summary}</button>
    <div id="{popover_id}" popover="auto" class="info-panel survey-info-popover">
      <button
        type="button"
        class="survey-info-close"
        popovertarget="{popover_id}"
        popovertargetaction="hide"
        aria-label="Close information panel"
      >Close<span class="survey-info-close-x" aria-hidden="true">×</span></button>
      <p><strong>{title}</strong></p>
      <p>Coal-to-clean projects seek to close a coal-fired power plant earlier than it would otherwise be expected to retire.</p>
      <p>Project finance may support early closure and decommissioning, new renewable electricity generation, grid-related measures, and support for affected workers and communities.</p>
      <p>The emissions avoided through early closure are compared with a baseline representing how long the plant would otherwise have operated. Verified net emission reductions before the baseline retirement date may be issued as carbon credits, sometimes called transition credits.</p>
      <p>Verra and Gold Standard have published methodologies for this project type. Several proposed transactions are being assessed or piloted, but the market remains at an early stage.</p>
      <p><a href="https://verra.org/methodologies/vm0052-accelerated-retirement-of-coal-fired-power-plants-using-a-just-transition-v1-0/" target="_blank" rel="noopener noreferrer">VM0052 Accelerated Retirement of Coal-Fired Power Plants Using a Just Transition</a></p>
      <p><a href="https://globalgoals.goldstandard.org/459_paa-m400-05_just-coal-decommissioning/" target="_blank" rel="noopener noreferrer">Gold Standard JUST: Coal Decommissioning methodology</a></p>
    </div>
    """


def validation_more_info(*, compact: bool = False) -> None:
    st.markdown(validation_more_info_html(compact=compact), unsafe_allow_html=True)


def render_survey_popup_behaviour() -> None:
    """Use native auto-popovers, which close on outside tap and Escape."""
    return


def post_validation_backend(
    action: str,
    payload: dict[str, Any],
    *,
    timeout: int,
) -> tuple[bool, str]:
    """Post survey data with one safe retry for an Apps Script cold-start timeout.

    The submission ID makes the retry idempotent in the research backend, so a
    response that reached Google just before the client timed out is not stored twice.
    """
    ok, message = post_backend(action, payload, timeout=timeout)
    if ok:
        return True, message

    lower_message = str(message).lower()
    if "timed out" not in lower_message and "timeout" not in lower_message:
        return False, message

    time.sleep(1.5)
    return post_backend(action, payload, timeout=timeout)


def register_validation_participant(profile: dict[str, Any]) -> tuple[bool, str]:
    return post_validation_backend(
        "log_validation_participant",
        {
            "participant_code": st.session_state.validation_participant_code,
            "registered_at": singapore_now(),
            "survey_version": VALIDATION_SURVEY_VERSION,
            "submission_id": st.session_state.validation_submission_id,
            **profile,
        },
        timeout=60,
    )


def submit_validation_responses(responses: list[dict[str, Any]]) -> tuple[bool, str]:
    return post_validation_backend(
        "log_validation_responses",
        {
            "participant_code": st.session_state.validation_participant_code,
            "submitted_at": singapore_now(),
            "survey_version": VALIDATION_SURVEY_VERSION,
            "submission_id": st.session_state.validation_submission_id,
            "responses": responses,
        },
        timeout=90,
    )


def survey_response_row(
    *,
    display_order: int,
    question_id: str,
    question_type: str,
    question: str,
    construct_id: str = "",
    construct: str = "",
    indicator_id: str = "",
    option_code: str = "",
    response_numeric: int | float | None = None,
    response_label: str = "",
    response_text: str = "",
    option_values: str = "",
) -> dict[str, Any]:
    return {
        "display_order": display_order,
        "question_id": question_id,
        "question_type": question_type,
        "construct_id": construct_id,
        "construct": construct,
        "indicator_id": indicator_id,
        "question": question,
        "option_code": option_code,
        "response_numeric": response_numeric,
        "response_label": response_label,
        "response_text": response_text,
        "option_values": option_values,
    }


def render_validation(index_data: pd.DataFrame) -> None:
    industries = sorted((name for name in INDUSTRY_MULTIPLIERS if name != "Other"), key=str.casefold)
    industry_options = ["Select your industry", *industries, "Other"]
    role_options = ["Select your role", *ROLE_OPTIONS]
    market_options = [*SURVEY_COUNTRY_OPTIONS, "Global or multi-country role", "Other"]
    # Offer the full jurisdiction list in straightforward alphabetical order.
    # Unlike the personal-profile selectors, this decision question should not
    # prioritise Singapore or ASEAN jurisdictions.
    host_jurisdictions = sorted(SURVEY_COUNTRIES, key=str.casefold)

    with st.container(key="validation_section"):
        st.markdown(
            '<div class="section-title">Research survey on coal-to-clean projects with carbon credits</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            """
            <div class="survey-intro">
              <p>Coal-to-clean projects aim to finance the early retirement of coal-fired power plants and their replacement with new renewable electricity generation through the sale of carbon credits, sometimes called transition credits.</p>
              <p>This survey examines which features of a host jurisdiction – including its government, energy system, carbon market, financing conditions and social protections – most affect whether carbon-credit buyers and other market professionals would support such projects there.</p>
              <div class="survey-time">Estimated completion time: 8–10 minutes</div>
              <div class="survey-required-note">All survey questions are required unless labelled optional.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        validation_more_info()
        render_survey_popup_behaviour()

        if st.session_state.validation_submitted:
            st.success(st.session_state.validation_submission_message or "Thank you. Your responses have been recorded. You may now close this window.")
            return

        with st.container(key="validation_form_shell"):
            st.markdown('<div class="survey-subtitle">Your details</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2, gap="large")
            with col1:
                name = st.text_input("Name", key="validation_name")
            with col2:
                organisation = st.text_input("Organisation name", key="validation_organisation")

            col1, col2 = st.columns(2, gap="large")
            with col1:
                email = st.text_input("Email address (optional)", key="validation_email")
            with col2:
                linkedin = st.text_input("LinkedIn profile (optional)", key="validation_linkedin")

            st.markdown(
                """
                <div class="survey-privacy survey-privacy-before-consent">Your details will be kept confidential and are requested solely for research purposes – to categorise respondent roles and, where you consent, to follow up about your responses.</div>
                """,
                unsafe_allow_html=True,
            )
            with st.container(key="validation_contact_consent_row"):
                contact_consent = st.checkbox(
                    "I consent to being contacted about my responses.",
                    key="validation_contact_consent",
                )
            st.markdown(
                """
                <div class="survey-privacy survey-privacy-after-consent">Your details will not be disclosed outside the research, except through services used to securely store responses. Findings will be reported only in aggregate or de-identified form.</div>
                """,
                unsafe_allow_html=True,
            )

            col1, col2 = st.columns(2, gap="large")
            with col1:
                industry = st.selectbox("Your industry", industry_options, key="validation_industry")
                industry_other = ""
                if industry == "Other":
                    industry_other = st.text_input(
                        "Please specify your industry or sector",
                        key="validation_industry_other",
                    )
            with col2:
                role = st.selectbox("Your role", role_options, key="validation_role")
                role_other = ""
                if role == "Other":
                    role_other = st.text_input(
                        "Please specify your role",
                        key="validation_role_other",
                    )

            col1, col2 = st.columns(2, gap="large")
            with col1:
                work_base = st.selectbox(
                    "In which country are you currently based for work?",
                    ["Select a country", *SURVEY_COUNTRY_OPTIONS],
                    key="validation_work_base",
                    on_change=clear_country_separator_from_selectbox,
                    args=("validation_work_base",),
                )
            with col2:
                main_markets = st.multiselect(
                    "Which countries do you work in, work with or follow most closely?",
                    market_options,
                    max_selections=5,
                    key="validation_main_markets",
                    help="Select up to five.",
                    on_change=clear_country_separator_from_multiselect,
                    args=("validation_main_markets",),
                )
                main_markets = [
                    value for value in main_markets if value != COUNTRY_GROUP_SEPARATOR
                ]
                markets_other = ""
                if "Other" in main_markets:
                    markets_other = st.text_input(
                        "Please specify the other market or markets",
                        key="validation_markets_other",
                    )

            st.markdown('<div class="survey-details-gap"></div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="survey-question-intro">When assessing a coal-to-clean opportunity, how important are the following?</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="survey-scenario-copy">
                  <div class="survey-scenario-lead"><span>Imagine that you are assessing whether to support a coal-to-clean transition-credit project.</span>&nbsp;{validation_more_info_html(compact=True)}</div>
                  <p>Support could involve buying credits through an offtake agreement, investing, financing, developing the project or recommending it professionally.</p>
                  <p>Please answer from your professional perspective, even if you do not personally purchase carbon credits or work directly with carbon projects.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            factor_responses: dict[str, int] = {}
            factor_question_numbers: dict[str, int] = {}
            for factor_number, item in enumerate(VALIDATION_FACTORS, start=1):
                factor_question_numbers[item["key"]] = factor_number
                st.markdown(
                    f'<div class="priority-label">Q{factor_number}. {html.escape(item["label"])}</div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div class="priority-description">{html.escape(item["description"])}</div>',
                    unsafe_allow_html=True,
                )
                slider_key = f'validation_factor_v5_{item["key"]}'
                with st.container(key=f'validation_factor_scale_{item["key"]}'):
                    st.markdown(
                        '<div class="survey-scale-label-row"><span>1 = Not important</span><span>5 = Critical</span></div>',
                        unsafe_allow_html=True,
                    )
                    factor_responses[item["key"]] = st.select_slider(
                        item["label"],
                        options=[0, 1, 2, 3, 4, 5],
                        value=0,
                        format_func=lambda value: "" if value == 0 else str(value),
                        key=slider_key,
                        label_visibility="collapsed",
                    )

            # Hide Streamlit's built-in lower endpoint labels. The selected numeric
            # value above the thumb remains visible; the unanswered sentinel is blank.
            components.html(
                """
                <script>
                (() => {
                  const doc = window.parent.document;
                  let cleaning = false;

                  function cleanSurveySliderLabels() {
                    if (cleaning) return;
                    cleaning = true;
                    try {
                      doc.querySelectorAll('[class*="st-key-validation_factor_scale_"] [data-testid="stSlider"]').forEach((slider) => {
                        const thumb = slider.querySelector('[role="slider"]');
                        if (!thumb) return;
                        const thumbRect = thumb.getBoundingClientRect();
                        const thumbCentreY = thumbRect.top + thumbRect.height / 2;

                        slider.querySelectorAll('.survey-hide-slider-label').forEach((el) => {
                          el.classList.remove('survey-hide-slider-label');
                        });

                        Array.from(slider.querySelectorAll('div, span, p')).forEach((el) => {
                          if (el.children.length !== 0) return;
                          const value = (el.textContent || '').trim();
                          if (!['0', '1', '2', '3', '4', '5', 'Select'].includes(value)) return;
                          const rect = el.getBoundingClientRect();
                          const centreY = rect.top + rect.height / 2;
                          const isLowerEndpoint = centreY > thumbCentreY + 5;
                          if (value === 'Select' || value === '0' || isLowerEndpoint) {
                            el.classList.add('survey-hide-slider-label');
                          }
                        });
                    } finally {
                      cleaning = false;
                    }
                  }

                  cleanSurveySliderLabels();
                  const observer = new MutationObserver(cleanSurveySliderLabels);
                  observer.observe(doc.body, {subtree: true, childList: true, characterData: true});
                  function enforceSurveySelectFonts() {
                    doc.querySelectorAll('[data-baseweb="popover"], [data-baseweb="select"], [role="listbox"]').forEach((node) => {
                      node.style.fontFamily = "Montserrat, sans-serif";
                      node.querySelectorAll('*').forEach((child) => {
                        child.style.fontFamily = "Montserrat, sans-serif";
                      });
                    });
                  }

                  enforceSurveySelectFonts();
                  const fontObserver = new MutationObserver(enforceSurveySelectFonts);
                  fontObserver.observe(doc.body, {subtree: true, childList: true});
                  window.addEventListener('beforeunload', () => {
                    observer.disconnect();
                    fontObserver.disconnect();
                  }, {once: true});
                })();
                </script>
                """,
                height=0,
            )

            st.markdown('<div class="survey-subtitle survey-later-question">Q23. Project’s host jurisdiction</div>', unsafe_allow_html=True)
            project_setting = st.radio(
                "All else being equal, in which type of jurisdiction would you be more inclined to support a coal-to-clean transition-credit project?",
                [
                    "Higher-income developed jurisdictions",
                    "Emerging and developing jurisdictions",
                    "No general preference",
                    "It depends on the particular project and jurisdiction",
                    "I do not have enough information to assess this",
                ],
                index=None,
                key="validation_project_setting",
            )

            st.markdown('<div class="survey-subtitle">Q24. Carbon-credit price for a coal-to-clean project</div>', unsafe_allow_html=True)
            price_response = st.radio(
                "Assuming the project meets robust integrity requirements, what is the highest price per tCO₂e you consider commercially plausible for its carbon credits?",
                [
                    "Below US$10",
                    "US$10–24",
                    "US$25–49",
                    "US$50–74",
                    "US$75–99",
                    "US$100–149",
                    "US$150 or more",
                    "It would depend too heavily on the project",
                    "I do not have enough information to assess this",
                ],
                index=None,
                key="validation_price",
            )

            st.markdown('<div class="survey-subtitle">Q25. Preferred host jurisdictions</div>', unsafe_allow_html=True)
            preferred_host_options = [
                *host_jurisdictions,
                "No particular jurisdiction preference",
                "I do not have enough information to assess this",
            ]
            preferred_hosts = st.multiselect(
                "All else being equal, which host jurisdictions would you be most willing to support through a five-year commercial commitment? (Choose up to 5)",
                preferred_host_options,
                max_selections=5,
                key="validation_preferred_hosts",
                on_change=clear_country_separator_from_multiselect,
                args=("validation_preferred_hosts",),
            )

            st.markdown('<div class="survey-subtitle">Q26. Coal phase-out initiatives</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="survey-help">Which initiatives do you believe have the greatest potential to accelerate coal phase-out? Rank up to three initiatives you know enough about to assess. Stop whenever you have no further choices.</div>',
                unsafe_allow_html=True,
            )
            unsure_initiative = "I do not know enough about these initiatives to rank them"
            first_choice_options = ["No selection", *COAL_PHASEOUT_INITIATIVES]
            later_choice_options = [
                "No further choice",
                *[
                    initiative
                    for initiative in COAL_PHASEOUT_INITIATIVES
                    if initiative != unsure_initiative
                ],
            ]
            rank1, rank2, rank3 = st.columns(3, gap="large")
            with rank1:
                initiative_1 = st.selectbox(
                    "First choice",
                    first_choice_options,
                    key="validation_initiative_1",
                )
            with rank2:
                initiative_2 = st.selectbox(
                    "Second choice",
                    later_choice_options,
                    key="validation_initiative_2",
                )
            with rank3:
                initiative_3 = st.selectbox(
                    "Third choice",
                    later_choice_options,
                    key="validation_initiative_3",
                )
            initiative_other = ""
            if "Other" in {initiative_1, initiative_2, initiative_3}:
                initiative_other = st.text_input(
                    "Please specify the other initiative",
                    key="validation_initiative_other",
                )

            st.markdown('<div class="survey-subtitle">Q27. Five most important factors</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="survey-help">Which five factors would have the greatest influence on your willingness to support a coal-to-clean transition-credit project?</div>',
                unsafe_allow_html=True,
            )
            top_state_keys = {
                item["key"]: f'validation_top_five_{item["key"]}'
                for item in VALIDATION_FACTORS
            }
            selected_before_render = {
                item["key"]
                for item in VALIDATION_FACTORS
                if bool(st.session_state.get(top_state_keys[item["key"]], False))
            }
            with st.container(key="validation_top_five_grid"):
                top_columns = st.columns(2, gap="large")
                for item_index, item in enumerate(VALIDATION_FACTORS):
                    state_key = top_state_keys[item["key"]]
                    already_selected = item["key"] in selected_before_render
                    disable_unselected = len(selected_before_render) >= 5 and not already_selected
                    with top_columns[item_index % 2]:
                        st.checkbox(
                            item["label"],
                            key=state_key,
                            disabled=disable_unselected,
                        )
            top_five = [
                item["label"]
                for item in VALIDATION_FACTORS
                if bool(st.session_state.get(top_state_keys[item["key"]], False))
            ]
            st.markdown(
                f'<div class="survey-counter">Selected: {len(top_five)} of 5</div>',
                unsafe_allow_html=True,
            )

            st.markdown('<div class="survey-subtitle">Q28. Anything to add?</div>', unsafe_allow_html=True)
            missing_factor = st.text_area(
                "Is there anything else about a jurisdiction, its government or its market conditions that would make you more or less willing to support a coal-to-clean transition-credit project there? (Optional)",
                key="validation_missing_factor",
            )

            with st.container(key="survey_submit_row"):
                submitted = st.button("Submit responses", type="primary", key="validation_submit", width="content")

            if not submitted:
                return

            errors: list[str] = []
            if not name.strip():
                errors.append("Name: enter your name.")
            if not organisation.strip():
                errors.append("Organisation name: enter your organisation name.")
            if contact_consent and not email.strip():
                errors.append("Email address: enter an email address because you consented to follow-up contact.")
            elif contact_consent and not valid_email(email):
                errors.append("Email address: enter a valid email address.")
            if industry == "Select your industry":
                errors.append("Your industry: select an industry.")
            if industry == "Other" and not industry_other.strip():
                errors.append("Your industry: specify your industry or sector.")
            if role == "Select your role":
                errors.append("Your role: select a role.")
            if role == "Other" and not role_other.strip():
                errors.append("Your role: specify your role.")
            if work_base == "Select a country":
                errors.append("Country of work base: select the country where you are currently based for work.")
            if not main_markets:
                errors.append("Main markets of experience: select at least one market.")
            if "Other" in main_markets and not markets_other.strip():
                errors.append("Main markets of experience: specify the other market or markets.")

            for item in VALIDATION_FACTORS:
                if int(factor_responses.get(item["key"], 0)) == 0:
                    question_number = factor_question_numbers[item["key"]]
                    errors.append(
                        f"Q{question_number} – {item['label']}: select an importance rating."
                    )

            if project_setting is None:
                errors.append("Q23 – Project’s host jurisdiction: select one response.")
            if price_response is None:
                errors.append("Q24 – Carbon-credit price for a coal-to-clean project: select one response.")
            if not preferred_hosts:
                errors.append("Q25 – Preferred host jurisdictions: select at least one response.")
            exclusive_host_answers = {
                "No particular jurisdiction preference",
                "I do not have enough information to assess this",
            }
            if any(value in preferred_hosts for value in exclusive_host_answers) and len(preferred_hosts) > 1:
                errors.append(
                    "Q25 – Preferred host jurisdictions: choose named jurisdictions or one no-preference option, not both."
                )
            initiative_choices = [
                value
                for value in [initiative_1, initiative_2, initiative_3]
                if value not in {"No selection", "No further choice"}
            ]
            if initiative_1 == "No selection":
                errors.append(
                    "Q26 – Coal phase-out initiatives: select a first choice, or indicate that you do not know enough to rank the initiatives."
                )
            if initiative_2 == "No further choice" and initiative_3 != "No further choice":
                errors.append(
                    "Q26 – Coal phase-out initiatives: choose a second-ranked initiative before adding a third-ranked initiative."
                )
            if initiative_1 == unsure_initiative and (
                initiative_2 != "No further choice"
                or initiative_3 != "No further choice"
            ):
                errors.append(
                    "Q26 – Coal phase-out initiatives: when selecting the insufficient-information response, leave the remaining choices blank."
                )
            if len(initiative_choices) != len(set(initiative_choices)):
                errors.append(
                    "Q26 – Coal phase-out initiatives: do not rank the same initiative more than once."
                )
            if "Other" in initiative_choices and not initiative_other.strip():
                errors.append(
                    "Q26 – Coal phase-out initiatives: specify the other initiative."
                )
            if len(top_five) != 5:
                errors.append("Q27 – Five most important factors: select exactly five factors.")

            if errors:
                error_items = "\n".join(f"- {error}" for error in errors)
                st.error(
                    "**Please complete the following required fields or questions:**\n\n" + error_items,
                    icon="⚠️",
                )
                return

            profile = {
                "name": name.strip(),
                "organisation": organisation.strip(),
                "email": email.strip(),
                "linkedin": linkedin.strip(),
                "consent_follow_up": "Yes" if contact_consent else "No",
                "industry": industry,
                "industry_other": industry_other.strip(),
                "role": role,
                "role_other": role_other.strip(),
                "work_base": work_base,
                "main_markets": " | ".join(main_markets),
                "main_markets_other": markets_other.strip(),
            }

            response_rows: list[dict[str, Any]] = []
            display_order = 1
            for item in VALIDATION_FACTORS:
                response_value = int(factor_responses[item["key"]])
                response_rows.append(
                    survey_response_row(
                        display_order=display_order,
                        question_id=item["question_id"],
                        question_type="importance_scale",
                        construct_id=item["construct_id"],
                        construct=item["construct"],
                        indicator_id=item["indicator_id"],
                        question=item["description"],
                        response_numeric=response_value,
                        response_label=VALIDATION_SCALE[response_value],
                        response_text=item["label"],
                        option_values="1=Not important | 2=Slightly important | 3=Moderately important | 4=Very important | 5=Critical",
                    )
                )
                display_order += 1

            response_rows.append(
                survey_response_row(
                    display_order=display_order,
                    question_id="C01",
                    question_type="single_choice",
                    question="All else being equal, in which type of jurisdiction would you be more inclined to support a coal-to-clean transition-credit project?",
                    option_code=project_setting,
                    response_label=project_setting,
                    option_values="Higher-income developed jurisdictions | Emerging and developing jurisdictions | No general preference | It depends on the particular project and jurisdiction | I do not have enough information to assess this",
                )
            )
            display_order += 1
            response_rows.append(
                survey_response_row(
                    display_order=display_order,
                    question_id="C02",
                    question_type="single_choice",
                    question="Highest commercially plausible price per tCO₂e",
                    option_code=price_response,
                    response_label=price_response,
                    option_values="Below US$10 | US$10–24 | US$25–49 | US$50–74 | US$75–99 | US$100–149 | US$150 or more | It would depend too heavily on the project | I do not have enough information to assess this",
                )
            )
            display_order += 1
            for host in preferred_hosts:
                response_rows.append(
                    survey_response_row(
                        display_order=display_order,
                        question_id="C03",
                        question_type="multi_choice",
                        question="Preferred host jurisdictions for a five-year commercial commitment",
                        option_code=host,
                        response_label=host,
                    )
                )
            display_order += 1
            for rank_number, initiative in enumerate(initiative_choices, start=1):
                response_rows.append(
                    survey_response_row(
                        display_order=display_order,
                        question_id="K01",
                        question_type="ranked_choice",
                        question="Coal phase-out initiatives with the greatest potential",
                        option_code=initiative,
                        response_numeric=rank_number,
                        response_label=f"Rank {rank_number}",
                        response_text=initiative_other.strip() if initiative == "Other" else initiative,
                    )
                )
            display_order += 1
            factor_by_label = {item["label"]: item for item in VALIDATION_FACTORS}
            for label in top_five:
                item = factor_by_label[label]
                response_rows.append(
                    survey_response_row(
                        display_order=display_order,
                        question_id="P01",
                        question_type="top_five",
                        construct_id=item["construct_id"],
                        construct=item["construct"],
                        indicator_id=item["indicator_id"],
                        question="Five factors with the greatest influence",
                        option_code=item["question_id"],
                        response_label=label,
                    )
                )
            display_order += 1
            if missing_factor.strip():
                response_rows.append(
                    survey_response_row(
                        display_order=display_order,
                        question_id="O01",
                        question_type="open_text",
                        question="Anything else about a jurisdiction, its government or market conditions that affects willingness to support a project",
                        response_text=missing_factor.strip(),
                    )
                )

            with st.spinner(
                "Submitting your responses. This may take up to a minute – please do not close or refresh this window.",
                show_time=True,
            ):
                participant_ok, participant_message = register_validation_participant(profile)
                if not participant_ok:
                    st.error("Your details were not recorded. " + participant_message)
                    return
                responses_ok, responses_message = submit_validation_responses(response_rows)
                if not responses_ok:
                    st.error("Your responses were not recorded. " + responses_message)
                    return

            st.session_state.validation_submitted = True
            # Keep direct-link respondents in survey mode after submission. This
            # avoids loading the desktop-oriented index on phones and prevents an
            # unexpected change of context immediately after the research task.
            st.session_state.survey_entry_locked = True
            st.session_state.validation_submission_message = (
                "Thank you. Your responses have been recorded. You may now close this window."
            )
            st.session_state.survey_success_pending = False
            st.success(st.session_state.validation_submission_message)
            return

METHODOLOGY_CONTENT = '\n<div class="methodology-copy">\n  <div class="methodology-subhead">Purpose and scope</div>\n  <p>The Coal-to-Clean Transition Jurisdiction Readiness Index is an initial screening tool for identifying jurisdictions with comparatively favourable conditions for credible coal-transition opportunities. It does not assess individual power plants or predict whether a specific transaction will succeed. Instead, it compares national-level energy, policy, institutional, carbon-market, financial and social conditions that could enable or constrain project development.</p>\n  <p>A market-confidence overlay supplements the structural index with quarterly evidence on changing investor, buyer, policy and implementation priorities. The overlay shows how recent market developments could alter the relative importance of existing readiness factors. It does not replace the structural base score.</p>\n  <p>Scores are comparative measures within the screened jurisdiction universe. They are not probabilities of project success and should not be interpreted as investment, legal or carbon-credit integrity advice.</p>\n\n  <div class="methodology-subhead">TRACTION and energy transition credits</div>\n  <p>The immediate conceptual starting point for this research was the Monetary Authority of Singapore-convened <a href="https://www.mas.gov.sg/development/sustainable-finance/transition-credits" target="_blank" rel="noopener noreferrer">Transition Credits Coalition (TRACTION)</a>. TRACTION was established to examine how carbon markets could complement existing financing mechanisms and accelerate Asia’s shift from coal-fired power to cleaner electricity while safeguarding energy reliability, affordability and access.</p>\n  <p>TRACTION proposed energy transition credits – carbon credits generated from verified emissions reductions achieved by retiring a coal-fired power plant earlier than its credible baseline retirement date, and replacing the lost generation with clean energy. The expected credit revenue can help bridge the financial gap created by foregone plant revenues, debt and contract obligations, renewable replacement costs, decommissioning and a Just Transition. Transition credits are therefore intended as one component of a wider financing package, not as a substitute for policy reform, power-system planning or conventional capital.</p>\n  <p>In the <a href="https://www.mas.gov.sg/-/media/mas-media-library/development/sustainable-finance/traction-final-report.pdf" target="_blank" rel="noopener noreferrer">TRACTION Final Report</a> released in November 2025, it provides a Selection and Prioritisation Framework that starts with the asset. Its Selection Criteria screen whether a coal-fired power plant can meet core integrity and feasibility requirements, first at a screening stage and then at pre-feasibility. Its Prioritisation Criteria then compare eligible opportunities at two levels. Asset-level criteria consider factors such as potential emissions reductions, job impacts, financial viability and the plant owner’s commitment. Market-level criteria assess four broad areas: dependence on coal; energy markets and policies; renewable-energy replacement prospects; and carbon-credit generation prospects.</p>\n  <p>The Coal-to-Clean Jurisdictional Readiness Index is an independent complement to that framework. It does not reproduce TRACTION’s asset-level eligibility test and cannot determine whether a particular plant can generate transition credits. Instead, it converts the market-level question into a reproducible cross-country screening tool. It separates opportunity materiality from readiness, compares 28 indicators across six pillars, and adds a quarterly market-confidence overlay. For market actors, it can help prioritise where to allocate asset-screening, engagement and due-diligence resources before undertaking the plant-level work recommended by TRACTION and required by formal crediting methodologies.</p>\n\n  <div class="methodology-subhead">Scope of coal use and methodological alignment</div>\n  <p>The index focuses on grid-connected coal-fired electricity generation. This is not a judgement that captive, off-grid or industrial uses of coal are unimportant. Grid-connected coal power provides the clearest basis for a reproducible jurisdictional comparison because plants are discrete, capacity-rated assets; electricity-system data are comparatively available across countries; and retirement can be assessed alongside replacement generation, grid reliability, power-sector policy and just-transition conditions.</p>\n  <p>The <a href="https://globalenergymonitor.org/projects/global-coal-plant-tracker/" target="_blank" rel="noopener noreferrer">Global Coal Plant Tracker</a> includes coal-fired electricity-generating units and separately identifies captive plants. For this index, operating units with a populated <code>Captive</code> field are excluded because GEM uses that field for power stations designated for particular non-grid use. Units without a captive designation are treated as the grid-connected comparison universe for index purposes, subject to manual review where other plant information indicates possible captive or mixed use.</p>\n  <p>This boundary aligns the index with the principal early coal-phaseout crediting methodologies. <a href="https://verra.org/methodologies/vm0052-accelerated-retirement-of-coal-fired-power-plants-using-a-just-transition-v1-0/" target="_blank" rel="noopener noreferrer">Verra VM0052</a> applies to accelerated retirement of grid-connected coal-fired power plants paired with renewable electricity replacement. <a href="https://globalgoals.goldstandard.org/459_paa-m400-05_just-coal-decommissioning/" target="_blank" rel="noopener noreferrer">Gold Standard’s JUST Coal Decommissioning methodology</a> similarly covers grid-connected coal plants and grid-connected renewable replacement.</p>\n  <p>Captive and industrial coal remain important transition challenges, but are not combined with the present index because their energy uses, replacement technologies, grid relationships and transaction structures differ substantially.</p>\n\n  <div class="methodology-subhead">Country universe and materiality screening</div>\n  <p>The index universe comprises jurisdictions with operating grid-connected coal-fired generation capacity recorded in Global Energy Monitor’s January 2026 Global Coal Plant Tracker. Applying the operational rule of Status = operating and a blank Captive field produces an eligible universe of 67 jurisdictions.</p>\n  <p>A central materiality screen is then applied across three dimensions:</p>\n  <ul>\n    <li>Operating grid-connected coal-fired generation capacity</li>\n    <li>Coal dependence within the electricity mix</li>\n    <li>Estimated lifetime emissions from the operating grid-connected coal fleet</li>\n  </ul>\n  <p>A jurisdiction passes where it meets at least two of the three median-based thresholds. Under the revised grid-connected universe, 34 jurisdictions pass the central screen. The screen measures the materiality of the coal-transition opportunity rather than readiness. Coal capacity and coal generation share are therefore not rewarded again within Pillar 1, avoiding double counting between opportunity scale and enabling conditions.</p>\n  \n\n  <div class="methodology-subhead">Index design</div>\n  <p>The index follows the principles set out in the <a href="https://www.oecd.org/en/publications/handbook-on-constructing-composite-indicators-methodology-and-user-guide_9789264043466-en.html" target="_blank" rel="noopener noreferrer">OECD and Joint Research Centre Handbook on Constructing Composite Indicators</a>. The design begins with a clear conceptual framework, selects indicators for relevance and comparability, normalises measures to a common scale, tests alternative weights, examines overlap and reports limitations. The index is intended to structure comparison and initiate deeper analysis, not to replace the underlying evidence or transaction-specific due diligence.</p>\n  <p>The 28 active indicators combine current conditions with recent trajectory measures. Current-condition indicators capture comparatively structural characteristics, such as renewable electricity penetration, governance quality and carbon-market experience. Trajectory indicators capture the direction of change, including recent trends in coal share, renewable electricity share and grid emissions intensity.</p>\n\n  <div class="methodology-subhead">Pillars and indicators</div>\n\n  <div class="pillar-heading">Pillar 1: Energy system conditions</div>\n  <p>This pillar assesses whether a power system already displays characteristics that can support coal replacement without treating coal dependence itself as readiness. For project developers, financiers, buyers and power-sector stakeholders, it provides an initial view of whether clean generation and grid decarbonisation are sufficiently established to support credible replacement planning, and where renewable integration or reliability risks may require deeper due diligence.</p>\n  <p>Operating coal capacity and coal generation share are retained as screening variables rather than scored indicators. This prevents large or highly coal-dependent systems from receiving a higher readiness score simply because the transition opportunity is larger.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Renewable electricity share.</strong> The share of national electricity generation supplied by renewable sources in 2024, measured as a percentage using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember’s Yearly Electricity Data</a>. A higher share indicates a stronger existing clean-generation base and greater operating experience with renewable power.</li>\n    <li><strong>Indicator 2: Wind and solar share.</strong> The combined share of national electricity generation supplied by wind and solar in 2024, measured as a percentage using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember</a>. This focuses on experience integrating variable renewable energy, which is particularly relevant to coal replacement and grid flexibility.</li>\n    <li><strong>Indicator 3: Grid emissions intensity.</strong> Average power-sector emissions in 2024, measured in grams of CO₂-equivalent per kilowatt-hour using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember</a>. Lower emissions intensity indicates a cleaner electricity system and less dependence on high-emitting generation.</li>\n  </ul>\n\n  <div class="pillar-heading">Pillar 2: Policy and transition commitment</div>\n  <p>This pillar examines whether government policy and regulation, coal-development decisions and recent power-sector trends are consistent with a durable transition. Developers, financiers and offtakers need more than an announced target: they need evidence that replacement-power investment is enabled, new coal development is constrained, and policy commitments are being implemented rather than leaving early retirement exposed to reversal, leakage or conflicting expansion plans.</p>\n  <p>Government action is central to the pillar, particularly through renewable-energy regulation, carbon-pricing infrastructure and dated phaseout commitments. It also looks beyond official announcements by using coal-project pipelines, operating-fleet alignment and observed electricity trends to capture the actions of utilities, asset owners and project sponsors. It does not directly score financial institutions’ coal-financing policies, which are not available on a sufficiently consistent cross-country basis.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Renewable energy policy and regulatory readiness.</strong> The country’s 2023 renewable-energy policy score from the World Bank’s <a href="https://www.worldbank.org/en/topic/energy/publication/rise---regulatory-indicators-for-sustainable-energy" target="_blank" rel="noopener noreferrer">Regulatory Indicators for Sustainable Energy</a>, expressed on a 0–100 scale. It reflects whether the policy and regulatory environment supports renewable deployment and investment.</li>\n    <li><strong>Indicator 2: Coal expansion pressure.</strong> The ratio of announced, pre-permit and permitted grid-connected coal capacity to operating grid-connected coal capacity in January 2026, calculated from the <a href="https://globalenergymonitor.org/projects/global-coal-plant-tracker/" target="_blank" rel="noopener noreferrer">Global Coal Plant Tracker</a>. A lower ratio indicates less forward pressure from the proposed coal pipeline.</li>\n    <li><strong>Indicator 3: Coal construction pressure.</strong> The ratio of grid-connected coal capacity under construction to operating grid-connected coal capacity in January 2026, calculated from <a href="https://globalenergymonitor.org/projects/global-coal-plant-tracker/" target="_blank" rel="noopener noreferrer">GEM</a>. It distinguishes projects already being built from earlier-stage proposals.</li>\n    <li><strong>Indicator 4: Coal share five-year trend.</strong> The linear change in coal’s share of electricity generation from 2019 to 2024, measured in percentage points per year using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember</a>. A declining trend indicates that coal is losing importance in the power mix.</li>\n    <li><strong>Indicator 5: Renewables share five-year trend.</strong> The linear change in renewable electricity share from 2019 to 2024, measured in percentage points per year using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember</a>. A rising trend provides evidence that clean generation is expanding rather than remaining static.</li>\n    <li><strong>Indicator 6: Grid emissions intensity five-year trend.</strong> The linear change in power-sector emissions intensity from 2019 to 2024, measured in grams of CO₂ per kilowatt-hour per year using <a href="https://ember-energy.org/data/yearly-electricity-data/" target="_blank" rel="noopener noreferrer">Ember</a>. A downward trend indicates that changes in the generation mix are translating into lower emissions.</li>\n    <li><strong>Indicator 7: Carbon-pricing infrastructure.</strong> A 0–100 score derived from the World Bank’s <a href="https://carbonpricingdashboard.worldbank.org/" target="_blank" rel="noopener noreferrer">Carbon Pricing Dashboard</a>. It considers implementation status, national or subnational scope, coverage of electricity and heat, explicit coal coverage, emissions coverage and the 2024 carbon price, while preventing early-stage or subnational instruments from receiving the same score as implemented national systems.</li>\n    <li><strong>Indicator 8: Coal phaseout commitment and documented implementation alignment.</strong> A 0–100 score based on whether a jurisdiction has a dated coal-phaseout target and the share of operating grid-connected coal capacity demonstrably aligned with that target. A net-zero year alone does not score because it does not establish a coal-specific implementation pathway.</li>\n  </ul>\n\n  <div class="pillar-heading">Pillar 3: Governance and institutional capacity</div>\n  <p>This pillar assesses whether public institutions and the legal environment can support credible project development and long-term commitments. Coal-phaseout and transition-credit transactions may require coordinated approvals, policy continuity, enforceable contracts, transparent public administration and confidence among asset owners, financiers and buyers that agreed retirement, replacement and credit-delivery arrangements will be upheld.</p>\n  <p>The five indicators use the World Bank’s <a href="https://www.worldbank.org/en/publication/worldwide-governance-indicators" target="_blank" rel="noopener noreferrer">Worldwide Governance Indicators</a>, converted to comparable 0–100 scores. They are broad national measures and do not replace project-specific legal or counterparty due diligence.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Political stability.</strong> The 2024 Political Stability and Absence of Violence measure. It provides a broad signal of disruption risk that could affect policy continuity, project implementation and long-term contractual arrangements.</li>\n    <li><strong>Indicator 2: Government effectiveness.</strong> The 2024 measure of public-service quality, policy formulation and implementation capacity. Stronger performance supports coordinated approvals and the delivery of complex transition programmes.</li>\n    <li><strong>Indicator 3: Regulatory quality.</strong> The 2024 measure of the government’s ability to formulate and implement policies and regulations that permit and promote private-sector development. It is relevant to renewable procurement, investment rules and carbon-market participation.</li>\n    <li><strong>Indicator 4: Rule of law.</strong> The 2024 measure of confidence in and adherence to the rules of society, including contract enforcement and property rights. It provides a high-level signal of whether transaction rights and obligations are likely to be enforceable.</li>\n    <li><strong>Indicator 5: Control of corruption.</strong> The 2024 measure of the extent to which public power is exercised for private gain. It is relevant to procurement integrity, permitting, public counterparties and confidence in the use of transition-related funds.</li>\n  </ul>\n\n  <div class="pillar-heading">Pillar 4: Carbon market maturity</div>\n  <p>This pillar assesses whether a jurisdiction has demonstrated experience with carbon-project development and whether the institutional infrastructure for international carbon cooperation is emerging. For transition-credit developers and financiers, this provides an initial signal of whether projects can move through validation and issuance, whether credits have reached end users, and whether host-country processes can support authorisation and transfer where required.</p>\n  <p>Voluntary carbon-market indicators capture demonstrated activity through 2024. Article 6 indicators capture more recent institutional and operational progress using the <a href="https://article6pipeline.unepccc.org/" target="_blank" rel="noopener noreferrer">UNEP Copenhagen Climate Centre Article 6 Pipeline</a>.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Voluntary carbon-market project activity.</strong> The all-time number of voluntary carbon projects hosted by the country through year-end 2024 in the <a href="https://gspp.berkeley.edu/berkeley-carbon-trading-project/offsets-database" target="_blank" rel="noopener noreferrer">Berkeley Voluntary Registry Offsets Database</a>. A larger project base indicates greater experience with project origination, validation and registry processes.</li>\n    <li><strong>Indicator 2: Voluntary carbon-market issuance scale.</strong> The all-time number of voluntary carbon credits issued from projects hosted by the country through year-end 2024, measured in tonnes of CO₂-equivalent using the <a href="https://gspp.berkeley.edu/berkeley-carbon-trading-project/offsets-database" target="_blank" rel="noopener noreferrer">Berkeley database</a>. It distinguishes project listings from demonstrated credit delivery.</li>\n    <li><strong>Indicator 3: Voluntary carbon-market retirement activity.</strong> The aggregate number of host-country credits retired from 2020 to 2024, measured in tonnes of CO₂-equivalent using the <a href="https://gspp.berkeley.edu/berkeley-carbon-trading-project/offsets-database" target="_blank" rel="noopener noreferrer">Berkeley database</a>. Retirements provide evidence of end-user demand and completed market use.</li>\n    <li><strong>Indicator 4: Article 6 institutional readiness.</strong> A 0–100 score based on five institutional and cooperation milestones recorded in the <a href="https://article6pipeline.unepccc.org/" target="_blank" rel="noopener noreferrer">Article 6 Pipeline</a>. It reflects whether the country has begun establishing the arrangements needed to participate credibly in cooperative carbon-market activity.</li>\n    <li><strong>Indicator 5: Article 6 operational maturity.</strong> A 0–100 stage score based on the most advanced implementation milestone recorded under Article 6.2 or the Paris Agreement Crediting Mechanism, ranging from an identified activity to recorded ITMO transfer or Article 6.4 emission-reduction issuance.</li>\n  </ul>\n\n  <div class="pillar-heading">Pillar 5: Macro-financial conditions</div>\n  <p>This pillar assesses whether the wider financial environment can support long-dated, capital-intensive coal-phaseout and transition-credit transactions. These transactions may need to finance early-retirement compensation, outstanding debt and contractual obligations, renewable replacement, decommissioning and Just Transition measures, while relying partly on future or cross-border carbon revenues. Even a technically credible project can struggle where local finance is shallow, inflation and exchange rates are unstable, or sovereign debt pressure constrains public support and raises the cost of capital.</p>\n  <p>The indicators are national macro-financial signals rather than project bankability tests. Market actors should use them to identify where financing structures, guarantees, currency hedging or concessional support may require greater attention.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Domestic financial depth.</strong> Domestic credit to the private sector in 2024, measured as a percentage of GDP using the World Bank’s <a href="https://databank.worldbank.org/source/world-development-indicators" target="_blank" rel="noopener noreferrer">World Development Indicators</a>. Greater financial depth suggests stronger capacity to intermediate capital and support private investment.</li>\n    <li><strong>Indicator 2: Macroeconomic price stability.</strong> Consumer-price inflation in 2024, measured as an annual percentage using the <a href="https://databank.worldbank.org/source/world-development-indicators" target="_blank" rel="noopener noreferrer">World Development Indicators</a>. Lower absolute inflation indicates a more predictable environment for costs, revenues and financing assumptions.</li>\n    <li><strong>Indicator 3: Sovereign debt pressure.</strong> General government gross debt in 2024, measured as a percentage of GDP using the International Monetary Fund’s <a href="https://www.imf.org/en/Publications/WEO/weo-database" target="_blank" rel="noopener noreferrer">World Economic Outlook Database</a>. Higher debt pressure may constrain fiscal support, guarantees and the government’s ability to absorb transition costs.</li>\n    <li><strong>Indicator 4: Exchange-rate instability.</strong> The average absolute annual logarithmic change in the official exchange rate from 2019 to 2024, calculated from the World Bank’s <a href="https://databank.worldbank.org/source/world-development-indicators" target="_blank" rel="noopener noreferrer">World Development Indicators</a> and expressed as a percentage. Greater instability increases currency-mismatch and repayment risk for cross-border financing and carbon revenues.</li>\n  </ul>\n\n  <div class="pillar-heading">Pillar 6: Just transition and social credibility</div>\n  <p>This pillar assesses whether the wider social and labour environment can support a fair and durable coal transition. Plant closure can affect workers, contractors, households, local businesses and public revenues, while credible transition-credit methodologies make Just Transition planning central to project integrity. Projects that do not recognise these impacts may face resistance, delays, reputational harm, buyer concern or reversal.</p>\n  <p>The indicators capture national enabling conditions rather than plant-level community consent or the adequacy of a particular Just Transition plan. For market actors, they identify where social dialogue, protection systems and labour-market absorption may require deeper assessment and dedicated financing.</p>\n  <ul class="indicator-list">\n    <li><strong>Indicator 1: Labour rights and social dialogue environment.</strong> A 0–100 score derived from the 2024 <a href="https://www.globalrightsindex.org/" target="_blank" rel="noopener noreferrer">ITUC Global Rights Index</a> rating, with a small adjustment for the recorded direction of change. It provides a broad signal of whether workers can organise, participate and raise concerns during transition planning.</li>\n    <li><strong>Indicator 2: Social protection and human-development resilience.</strong> A composite 0–100 score combining population coverage by social-protection systems from <a href="https://rshiny.ilo.org/dataexplorer20/" target="_blank" rel="noopener noreferrer">ILOSTAT</a> with the 2023 <a href="https://hdr.undp.org/data-center" target="_blank" rel="noopener noreferrer">UNDP Human Development Index</a>. It reflects the extent to which workers and communities have broader institutional and human-development buffers against transition disruption.</li>\n    <li><strong>Indicator 3: Labour-market absorption risk.</strong> Total unemployment in 2024, measured as a percentage of the labour force using the World Bank’s <a href="https://databank.worldbank.org/source/world-development-indicators" target="_blank" rel="noopener noreferrer">World Development Indicators</a>. Lower unemployment suggests greater capacity for displaced workers to find alternative employment, although local coal-region conditions may differ markedly from the national average.</li>\n  </ul>\n\n  <div class="methodology-subhead">Data period and annual update approach</div>\n  <p>The principal base year for the index is 2024, the latest year providing sufficiently comparable coverage across most indicators and jurisdictions. Exceptions are limited to essential sources that were unavailable for 2024 or represent current institutional or project status rather than annual outcomes.</p>\n  <p>The RISE renewable-energy policy indicator uses 2023 data. The Human Development Index also uses 2023 data. Certain social indicators use the closest complete observation to 2024 where full 2024 coverage was unavailable. Coal-fleet eligibility and project status use the January 2026 Global Coal Plant Tracker. Article 6 indicators use the UNEP-CCC data available at the disclosed analytical cut-off.</p>\n  <p>The 2024 indicators measure underlying readiness, while the more recent trackers define the current coal-transition opportunity set and capture institutional developments that can change between annual statistical releases. Each future edition should disclose data cut-offs and any material changes in indicator definition, source coverage or treatment.</p>\n\n  <div class="methodology-subhead">Data treatment and scoring</div>\n  <p>Raw indicators are aligned so that higher scores consistently represent stronger readiness. Measures of risk or transition pressure, including coal construction, inflation, sovereign debt pressure and exchange-rate instability, are reverse-scored.</p>\n  <p>Indicators are converted to a common 0–100 scale. Percentile bounds are used where necessary to reduce the influence of extreme outliers. Logarithmic transformations are applied to highly skewed carbon-market measures, including project counts, credit issuances and credit retirements. Indicator scores are aggregated into pillar scores, which are then combined into the overall index score.</p>\n  <p>Missing observations are not automatically treated as zero. Where a defensible estimate is required, it is documented in the analytical model and tested through sensitivity analysis. Where missing data remain, a pillar score requires valid observations covering at least 80% of its original indicator weight. Available weights are proportionately renormalised; the pillar score remains unavailable where the threshold is not met.</p>\n\n  <div class="methodology-subhead">Pillar weights</div>\n  <table class="methodology-weight-table">\n    <thead><tr><th>Pillar</th><th>Overall weight</th><th>Active indicators</th></tr></thead>\n    <tbody>\n      <tr><td>Energy system conditions</td><td>20%</td><td>3</td></tr>\n      <tr><td>Policy and transition commitment</td><td>20%</td><td>8</td></tr>\n      <tr><td>Governance and institutional capacity</td><td>15%</td><td>5</td></tr>\n      <tr><td>Carbon market maturity</td><td>20%</td><td>5</td></tr>\n      <tr><td>Macro-financial conditions</td><td>15%</td><td>4</td></tr>\n      <tr><td>Just transition and social credibility</td><td>10%</td><td>3</td></tr>\n      <tr><td><strong>Total</strong></td><td><strong>100%</strong></td><td><strong>28</strong></td></tr>\n    </tbody>\n  </table>\n  <p>The pillar weights reflect each pillar’s relevance to coal-transition execution, the reliability and coverage of its evidence, and the need to limit duplication across related risks. Indicators within each pillar are weighted according to their relationship with the pillar’s due-diligence question, data quality and distinctiveness from other measures. Detailed indicator weights are retained within the analytical model rather than reproduced on this page.</p>\n\n  <div class="methodology-subhead">Market-confidence overlay</div>\n  <p>The quarterly market-confidence overlay assesses which readiness factors received increased attention from investors, buyers, policymakers and other market participants during each quarter. It does not assign positive or negative media scores to individual jurisdictions. Instead, it adjusts the relative emphasis placed on existing indicators, with the same quarterly adjustments applied consistently across all jurisdictions.</p>\n  <p>Nexis is used as the main systematic discovery corpus, supported by targeted Google searches for primary-source verification. A Q1 pilot using QCIntel informed the development of the initial constructs, terminology and coding rules but is treated as a methodology-development corpus rather than part of the formal quarterly comparison.</p>\n  <p>Duplicate reporting of the same development is clustered into a single evidence event. Several linked pages and documents concerning the same underlying development are treated as a related source bundle unless they provide substantively different decision-relevant evidence. Opposing claims about the same construct within one source are coded as separate actor–claim–construct records so that actor, direction and mechanism are preserved; they remain one source and usually one evidence event.</p>\n  <p>Evidence is assessed according to its relevance to coal transition, evidence type, signal strength and source independence. Construct-level evidence is converted into multipliers of 1.00, 1.05, 1.10 or 1.20. Each scored construct is assigned to its closest primary index indicator. Cross-cutting or project-specific issues without a sufficiently direct country-level indicator are retained as contextual findings rather than forced into the scoring model. Adjusted weights are renormalised to 100% before quarterly scores are calculated.</p>\n\n  <div class="methodology-subhead">Evolution of transition-credit methodologies</div>\n  <p>The ability of transition-credit projects to scale will depend partly on the development and market acceptance of credible methodologies. <a href="https://verra.org/methodologies/vm0052-accelerated-retirement-of-coal-fired-power-plants-using-a-just-transition-v1-0/" target="_blank" rel="noopener noreferrer">Verra VM0052</a> and <a href="https://globalgoals.goldstandard.org/459_paa-m400-05_just-coal-decommissioning/" target="_blank" rel="noopener noreferrer">Gold Standard’s JUST Coal Decommissioning methodology</a> convert complex retirement transactions into auditable requirements covering baseline retirement dates, additionality, renewable replacement, monitoring and impacts on workers and communities.</p>\n  <p>These frameworks cannot remove all uncertainty. Results remain sensitive to the counterfactual retirement date, financial and regulatory additionality, replacement-power delivery, grid responses, leakage, plant operating assumptions and the enforceability of retirement commitments. Verra’s <a href="https://verra.org/methodologies/proposed-revision-to-vm0052-accelerated-retirement-of-coal-fired-power-plants-using-a-just-transition-v1-0/" target="_blank" rel="noopener noreferrer">proposed revision to VM0052</a> illustrates that baseline determination, replacement electricity and grid-emissions treatment remain areas of methodological development.</p>\n  <p>The index therefore applies disclosed annual and quarterly cut-off dates. Material changes in methodologies, guidance from the <a href="https://www.icvcm.org/continuous-improvement-work-programs/standardised-approaches/" target="_blank" rel="noopener noreferrer">Integrity Council for the Voluntary Carbon Market</a> or established market practice may inform subsequent editions or quarterly market-salience analysis, but do not automatically rewrite previously published scores.</p>\n\n  <div class="methodology-subhead">Validation and sensitivity testing</div>\n  <p>Preliminary correlation testing found a maximum inter-pillar correlation of 0.764, below the 0.800 threshold used to flag potentially excessive overlap. The pillar and indicator weighting scenarios are being rerun using the corrected sensitivity model, and final rank-correlation statistics will replace the preliminary results when testing is complete.</p>\n  <p>Final recoded overlay testing produced Spearman rank correlations of 0.998 for Q1 and 0.998 for Q2 against the base index, with no jurisdiction moving by more than two ranking positions. The limited movement is intentional: the overlay adds current market context without overpowering the slower-moving structural findings of the base index.</p>\n  <div class="validation-box">\n    <div class="validation-title">Preliminary robustness summary</div>\n    <div class="validation-grid">\n      <div>Alternative pillar and indicator weighting scenarios</div><div class="validation-value">Final rerun pending</div>\n      <div>Highest inter-pillar correlation</div><div class="validation-value">0.764</div>\n      <div>Base-to-Q1 and base-to-Q2 rank correlations</div><div class="validation-value">0.998 and 0.998</div>\n      <div>Maximum quarterly overlay movement</div><div class="validation-value">2 positions</div>\n    </div>\n  </div>\n\n  <div class="methodology-subhead">User-priority survey and external validation</div>\n  <p>Optional user-priority responses, together with industry and role classifications, are collected to compare stated market priorities with the constructs identified through quarterly evidence coding. Responses are analysed and reported only in aggregate and may be used to validate or refine subsequent overlays, annual methodology revisions or separately labelled sensitivity analyses.</p>\n  <p>Survey responses do not retrospectively alter a published overlay unless a pre-specified revision protocol, minimum response threshold and revised model version are disclosed. Personal contact details are used only to fulfil report-delivery requests and support the stated research purpose; they are not used for marketing.</p>\n  \n\n  <div class="methodology-subhead">AI-assisted analysis and quality control</div>\n  <p>Generative AI is used as an analytical aid rather than as an independent source of evidence. It supports data extraction, evidence coding, formula development, consistency checks and iterative analytical testing. The researcher designs and refines prompts, establishes the classification framework, reviews source documents and corrects coding or calculation decisions. AI-generated outputs are treated as provisional until checked against the underlying dataset, document or evidence source. The researcher retains responsibility for the final methodology, calculations and interpretation.</p>\n\n\n  <div class="methodology-subhead">Limitations</div>\n  <p>The index is a national-level screening tool. Coal-transition transactions ultimately depend on plant-level economics, ownership, contracts, grid access, technical feasibility and community consent. Some international datasets are published with substantial time lags, and English-language evidence availability is uneven across jurisdictions.</p>\n  <div class="pillar-heading">Four unresolved transition tensions</div>\n  <ul class="indicator-list">\n    <li><strong>Additionality versus readiness:</strong> Strong policy and improving economics can make retirement more feasible, but may weaken the case that carbon finance caused it; only a plant-level baseline can resolve this.</li>\n    <li><strong>Transition need versus implementation capability:</strong> Jurisdictions with the largest financing need and emissions opportunity may also have the weakest governance, financial or delivery conditions.</li>\n    <li><strong>Younger fleets versus retirement feasibility:</strong> Newer plants offer greater avoided carbon lock-in but often carry more debt and stronger contractual protection, while older plants may be easier to close but closer to retirement anyway.</li>\n    <li><strong>Speed versus reliability and justice:</strong> Faster closure can increase avoided emissions only where replacement power, grid upgrades and measures for workers and communities are delivered in step.</li>\n  </ul>\n  <p>These tensions remain after the revised non-captive screen because the index separates jurisdictional readiness from plant-level eligibility; it can show where trade-offs may be more manageable, but cannot resolve them for an individual transaction.</p>\n  <p>The index does not directly determine whether a jurisdiction has sufficient technically, spatially and economically feasible clean-energy resources to replace a particular coal fleet. National conditions can conceal limited land availability, weak solar or wind resources in relevant locations, transmission bottlenecks, storage requirements, permitting restrictions and the distance between renewable resources and demand centres.</p>\n  <p>Existing energy-system indicators partly capture actual progress in renewable deployment, but they do not constitute plant-level replacement-power studies. Future editions may consider comparable evidence on renewable-resource potential, grid expansion, storage, cross-border interconnection and clean-electricity imports. Any addition would need to distinguish deliverable replacement electricity from theoretical resource potential and avoid overlap with existing energy-system indicators.</p>\n  <p>Several other relevant issues cannot be measured consistently across the full jurisdiction universe, including plant-specific community support, local grid-connection constraints, contractual arrangements and transaction-level bankability. These factors are reserved for subsequent project due diligence rather than represented through weak or inconsistent national proxies.</p>\n  <p class="methodology-footer">COAL-TO-CLEAN JURISDICTIONAL READINESS INDEX 2026&nbsp;&nbsp;|&nbsp;&nbsp;DEVELOPED BY GRACE TAY, EMSC SUSTAINABILITY MANAGEMENT&nbsp;&nbsp;|&nbsp;&nbsp;CONTACT: K2521144H@E.NTU.EDU.SG</p>\n</div>\n'


def render_methodology() -> None:
    with st.container(key="methodology_section"):
        st.markdown('<div class="section-title">Methodology</div>', unsafe_allow_html=True)
        st.markdown(METHODOLOGY_CONTENT, unsafe_allow_html=True)


initialise_state()

try:
    index_data = load_index_data(str(DATA_FILE))
    overlay_data = load_market_overlay(str(DATA_FILE))
    market_drivers, market_context, market_developments = load_market_content(
    str(DATA_FILE),
    DATA_FILE.stat().st_mtime,)
except Exception as exc:
    st.error(str(exc))
    st.stop()

survey_route_active = query_value(SURVEY_ROUTE_PARAM, "") == SURVEY_ROUTE_TOKEN
if survey_route_active:
    st.session_state.survey_entry_locked = True

survey_mode_active = bool(
    survey_route_active
    or st.session_state.survey_entry_locked
)
view = "validation" if survey_mode_active else query_value("view", "overall")
if view not in {"overall", "market", "priorities", "methodology", "validation"}:
    view = "overall"
if view == "validation" and not survey_mode_active:
    view = "overall"

if view == "validation":
    st.markdown(
        "<style>[data-testid='stSidebar'] { display: none !important; }</style>",
        unsafe_allow_html=True,
    )
    render_survey_header()
else:
    render_sidebar(view)
    render_header()
    render_mobile_desktop_notice()

if view == "market":
    render_market(index_data, overlay_data, market_drivers, market_context, market_developments)
elif view == "priorities":
    render_priorities(index_data)
elif view == "methodology":
    render_methodology()
elif view == "validation":
    render_validation(index_data)
else:
    render_overall(index_data)

if st.session_state.scroll_to_top or st.session_state.scroll_target:
    target_id = st.session_state.scroll_target or "top"
    components.html(
        f"""
        <script>
        (() => {{
          const win = window.parent;
          const doc = win.document;
          const targetId = {json.dumps(target_id)};
          const scrollTarget = () => {{
            try {{
              if (doc.activeElement && typeof doc.activeElement.blur === 'function') {{
                doc.activeElement.blur();
              }}
              const candidates = [
                doc.querySelector('section[data-testid="stMain"]'),
                doc.querySelector('[data-testid="stMain"]'),
                doc.querySelector('[data-testid="stAppViewContainer"]'),
                doc.scrollingElement,
                doc.documentElement,
                doc.body
              ].filter(Boolean);
              if (targetId === 'top') {{
                candidates.forEach((target) => {{
                  target.scrollTop = 0;
                  if (typeof target.scrollTo === 'function') target.scrollTo(0, 0);
                }});
                win.scrollTo(0, 0);
              }}
              const anchor = doc.getElementById(targetId);
              if (anchor && typeof anchor.scrollIntoView === 'function') {{
                anchor.scrollIntoView({{block: 'start', inline: 'nearest'}});
              }}
              const url = new URL(win.location.href);
              url.hash = targetId;
              win.history.replaceState({{}}, '', url.toString());
            }} catch (error) {{
              win.scrollTo(0, 0);
            }}
          }};
          win.setTimeout(scrollTarget, 100);
        }})();
        </script>
        """,
        height=0,
        width=0,
    )
    st.session_state.scroll_to_top = False
    st.session_state.scroll_target = None

if st.session_state.gate_action == "download":
    download_gate_dialog(index_data)
elif st.session_state.gate_action == "email":
    email_gate_dialog(index_data)
